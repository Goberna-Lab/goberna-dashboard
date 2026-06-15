import csv
import json
import logging
import os
import urllib.error
import urllib.parse
import urllib.request
from decimal import Decimal
from io import StringIO, BytesIO
import datetime
import re
import unicodedata
from difflib import SequenceMatcher
from urllib.parse import urljoin

from django.contrib.auth.decorators import login_required
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import (
    Case,
    Count,
    DateTimeField,
    DecimalField,
    ExpressionWrapper,
    F,
    Q as DJANGO_Q,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Coalesce, ExtractYear, ExtractMonth, TruncMonth
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import render, redirect
from django.conf import settings
from django.core.cache import cache
from django.utils import timezone
import hashlib

from .models import Venta, Cuota, Moneda, PerfilUsuario, DetalleVenta, LibroEnPack, MetaAds, MetaAccount, MetaCampaignMap, Producto

try:
    import openpyxl
except ImportError:
    openpyxl = None


logger = logging.getLogger(__name__)

DASHBOARD_GLOBAL_USER_IDS = {7, 8, 35}
ADMIN_GROUP_IDS = (2,)
DASHBOARD_SCOPE_GROUP_NAMES = ("Scope - Dashboard Satelite Global",)

# User decision: only 2026 campaigns require linking — 2025-only historic campaigns
# remain in the Pauta data but are NOT counted as pending anywhere in the UI.
LINKING_SCOPE_FROM = datetime.date(2026, 1, 1)
MEDIO_LABELS = {
    "organico": "Orgánico",
    "pagado": "Pagado",
    "referente": "Referente",
    "remarketing": "Remarketing",
    "postventa": "Postventa",
}


def _is_admin_user(user) -> bool:
    if not user or not getattr(user, "is_authenticated", False):
        return False
    return bool(
        user.is_superuser
        or user.id in DASHBOARD_GLOBAL_USER_IDS
        or user.groups.filter(id__in=ADMIN_GROUP_IDS).exists()
        or user.groups.filter(name__in=DASHBOARD_SCOPE_GROUP_NAMES).exists()
    )


def _medio_label(value: str) -> str:
    key = (value or "").strip().lower()
    if not key:
        return "Sin medio"
    return MEDIO_LABELS.get(key, key.replace("_", " ").title())


def _resolve_media_url(path_value: str) -> str:
    """Normaliza rutas de media guardadas en la BD principal a una URL absoluta usable en Vercel."""
    path = (path_value or "").strip()
    if not path:
        path = "productos/default.png"

    if path.startswith(("http://", "https://")):
        return path

    main_app_url = getattr(settings, "MAIN_APP_URL", "https://app.goberna.pe").rstrip("/") + "/"
    media_url = getattr(settings, "MEDIA_URL", "/media/")

    if path.startswith("/"):
        return urljoin(main_app_url, path.lstrip("/"))

    if isinstance(media_url, str) and media_url.startswith(("http://", "https://")):
        media_base = media_url if media_url.endswith("/") else f"{media_url}/"
    else:
        media_path = str(media_url or "/media/")
        if not media_path.startswith("/"):
            media_path = f"/{media_path}"
        media_base = f"{main_app_url.rstrip('/')}{media_path}"
        if not media_base.endswith("/"):
            media_base = f"{media_base}/"

    return urljoin(media_base, path.lstrip("/"))

def _is_local_request(request) -> bool:
    """Solo hosts de desarrollo local. Defensa en profundidad: aunque DEBUG=True
    quede mal seteado en producción, el bypass de login JAMÁS aplica fuera de localhost."""
    host = request.get_host().split(":")[0]
    return host in ("127.0.0.1", "localhost")


def login_required_unless_debug(view_func):
    """En producción exige login siempre. El bypass solo existe para preview local
    (DEBUG=True Y host localhost) por la cookie de sesión cross-domain de .goberna.pe."""
    decorated = login_required(view_func)

    def _wrapped(request, *args, **kwargs):
        if settings.DEBUG and _is_local_request(request):
            return view_func(request, *args, **kwargs)
        return decorated(request, *args, **kwargs)

    _wrapped.__name__ = getattr(view_func, "__name__", "_wrapped")
    return _wrapped


def ads_admin_required(view_func):
    """
    Protege las pantallas de Ads/Pauta: solo usuarios admin pueden acceder.
    - Preview local (DEBUG=True + localhost): se permite sin login.
    - Producción sin sesión: redirige al LOGIN_URL (igual que login_required).
    - Autenticado pero no-admin: 403 con página estilizada en español.
    """
    decorated = login_required(view_func)

    def _wrapped(request, *args, **kwargs):
        # Preview local sin login — mismo bypass que login_required_unless_debug
        if settings.DEBUG and _is_local_request(request):
            return view_func(request, *args, **kwargs)
        # En producción: primero exigir login
        if not request.user.is_authenticated:
            return decorated(request, *args, **kwargs)
        # Autenticado pero sin permiso admin
        if not _is_admin_user(request.user):
            return HttpResponseForbidden(
                """<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Sin acceso — Goberna</title>
  <style>
    body { font-family: 'Segoe UI', sans-serif; background: #f0f4f8; display: flex; align-items: center; justify-content: center; min-height: 100vh; margin: 0; }
    .card { background: #fff; border-radius: 12px; box-shadow: 0 4px 24px rgba(0,0,0,.10); padding: 48px 40px; max-width: 420px; text-align: center; }
    .icon { font-size: 3rem; color: #e53e3e; margin-bottom: 16px; }
    h1 { color: #1f2d3d; font-size: 1.4rem; margin: 0 0 12px; }
    p { color: #4a5568; font-size: 0.97rem; line-height: 1.6; margin: 0 0 24px; }
    a { display: inline-block; background: #1f2d3d; color: #fff; padding: 10px 24px; border-radius: 8px; text-decoration: none; font-weight: 600; }
    a:hover { background: #2d3f52; }
  </style>
</head>
<body>
  <div class="card">
    <div class="icon">&#128274;</div>
    <h1>Acceso restringido</h1>
    <p>No tenés permisos para ver la sección de Ads / Pauta.<br>
       Esta área es exclusiva para administradores.</p>
    <a href="/">Volver al inicio</a>
  </div>
</body>
</html>"""
            )
        return view_func(request, *args, **kwargs)

    _wrapped.__name__ = getattr(view_func, "__name__", "_wrapped")
    return _wrapped


@login_required_unless_debug
def home_dashboard(request):
    """Dashboard de ventas por usuario actual con opción de exportar reportes."""
    # Preview local sin sesión → vista global de admin (solo localhost + DEBUG).
    is_admin = _is_admin_user(request.user) or (
        settings.DEBUG and not request.user.is_authenticated and _is_local_request(request)
    )
    # cambios del vvito
    # Base scope por usuario
    if is_admin:
        ventas_scope = Venta.objects.all()
    else:
        ventas_scope = Venta.objects.filter(usuario=request.user)
    cuotas_scope = Cuota.objects.filter(venta__in=ventas_scope)

    # Universo de ventas del dashboard:
    # - Pagadas (1) y Pendientes (2)..
    # - Sin exigir confirmación de Tesorería para estado=1.
    ventas_base = ventas_scope.filter(estado__in=[1, 2])
    cuotas_base = Cuota.objects.filter(venta__in=ventas_base)

    # CÁLCULO EN DÓLARES (USD)
    
    # Anotar Ventas con monto_usd
    ventas_qs = ventas_base.annotate(
        tasa_cambio=Coalesce(
            'radio_multiplicador_usado', 
            'moneda__radioMultiplicador', 
            1,
            output_field=DecimalField()
        )
    ).annotate(
        tasa_final=Case(
            When(tasa_cambio=0, then=Value(1)),
            default=F('tasa_cambio'),
            output_field=DecimalField()
        )
    ).annotate(
        monto_usd=ExpressionWrapper(
            F('monto_total') / F('tasa_final'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    ).annotate(
        # Regla de fecha del dashboard:
        # - Siempre usar fecha_venta.
        # - Fallback a fecha_registro solo si fecha_venta viene nula.
        fecha_evento=Coalesce("fecha_venta", "fecha_registro", output_field=DateTimeField())
    )

    # Anotar Cuotas con monto_usd
    cuotas_qs = cuotas_base.select_related('venta', 'venta__moneda').annotate(
        tasa_cambio=Coalesce(
            'venta__radio_multiplicador_usado', 
            'venta__moneda__radioMultiplicador', 
            1,
            output_field=DecimalField()
        )
    ).annotate(
        tasa_final=Case(
            When(tasa_cambio=0, then=Value(1)),
            default=F('tasa_cambio'),
            output_field=DecimalField()
        )
    ).annotate(
        monto_usd=ExpressionWrapper(
            F('monto_total') / F('tasa_final'),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    )

    # Exportaciones
    reporte = request.GET.get("reporte")
    fmt = request.GET.get("fmt", "csv").lower()
    if reporte == "ventas":
        return _export_ventas(ventas_qs, fmt)
    if reporte == "cuotas":
        return _export_cuotas(cuotas_qs, fmt)

    # CACHÉ DE ESTADÍSTICAS
    params_sorted = sorted(request.GET.items())
    params_hash = hashlib.md5(str(params_sorted).encode()).hexdigest()
    # v12: ranking de vendedores total = deuda + pagado (cuotas USD)
    cache_key = f"dash_stats_v12_pending_paid_{request.user.id}_{is_admin}_{params_hash}"
    cached_stats = cache.get(cache_key)

    if cached_stats:
        ventas_resumen = cached_stats["ventas_resumen"]
        ventas_por_estado = cached_stats["ventas_por_estado"]
        monthly = cached_stats["monthly"]
        cuotas_por_estado = cached_stats["cuotas_por_estado"]
        ventas_chart_labels = cached_stats["ventas_chart_labels"]
        ventas_chart_data = cached_stats["ventas_chart_data"]
        ventas_estado_chart_data = cached_stats["ventas_estado_chart_data"]
        cuotas_estado_chart_data = cached_stats["cuotas_estado_chart_data"]
        cat_labels = cached_stats.get("cat_labels", [])
        cat_data = cached_stats.get("cat_data", [])
        books_labels = cached_stats.get("books_labels", [])
        books_data = cached_stats.get("books_data", [])
        courses_labels = cached_stats.get("courses_labels", [])
        courses_data = cached_stats.get("courses_data", [])
        vendors_data = cached_stats.get("vendors_data", [])
        country_labels = cached_stats.get("country_labels", [])
        country_data = cached_stats.get("country_data", [])
        medium_labels = cached_stats.get("medium_labels", [])
        medium_data = cached_stats.get("medium_data", [])
    else:
        # CALCULAR
        try:
            ventas_resumen = ventas_qs.aggregate(
                total=Count("id"),
                monto=Sum(ExpressionWrapper(
                    F('monto_total') / Case(
                        When(moneda__radioMultiplicador__isnull=True, then=Value(1)),
                        When(moneda__radioMultiplicador=0, then=Value(1)),
                        default=F('moneda__radioMultiplicador'),
                        output_field=DecimalField()
                    ),
                    output_field=DecimalField()
                ))
            )
        except Exception:
            ventas_resumen = {"total": 0, "monto": 0}

        # Gráficos (Mensual)
        monthly_qs = ventas_qs.annotate(
            year=ExtractYear('fecha_evento'),
            month=ExtractMonth('fecha_evento')
        ).values('year', 'month').annotate(
            total_usd=Sum(ExpressionWrapper(
                F('monto_total') / Case(
                    When(moneda__radioMultiplicador__isnull=True, then=Value(1)),
                    When(moneda__radioMultiplicador=0, then=Value(1)),
                    default=F('moneda__radioMultiplicador'),
                    output_field=DecimalField()
                ),
                output_field=DecimalField()
            )),
            count=Count('id')
        ).order_by('year', 'month')

        monthly = []
        for m in monthly_qs:
            try:
                y = m['year'] or 2024
                mo = m['month'] or 1
                date_obj = datetime.date(y, mo, 1)
                m_iso = date_obj.isoformat()
            except:
                m_iso = "Unknown"
            monthly.append({
                "month": m_iso,
                "total": m['total_usd'] or 0,
                "count": m['count']
            })

        # Estados
        ventas_estado_qs = ventas_scope.values('estado').annotate(total=Count('id'))
        ventas_estado_keys = [1, 2, 3, 4, 5, 6, 7]
        ventas_estado_map = {int(v["estado"]): int(v["total"] or 0) for v in ventas_estado_qs}
        ventas_estado_chart_data = [ventas_estado_map.get(k, 0) for k in ventas_estado_keys]
        
        ventas_por_estado = [] 

        cuotas_estado_qs = cuotas_scope.values('estado').annotate(total=Count('id'))
        cuotas_estado_keys = [1, 2, 3, 4, 5]
        cuotas_estado_map = {int(c["estado"]): int(c["total"] or 0) for c in cuotas_estado_qs}
        cuotas_estado_chart_data = [cuotas_estado_map.get(k, 0) for k in cuotas_estado_keys]
        cuotas_por_estado = []

        # Categorías
        detalles_qs = DetalleVenta.objects.filter(venta__in=ventas_qs).select_related(
            'venta', 'venta__moneda', 'producto__codigo_categoria', 'producto__codigo_negocio'
        ).annotate(
            tasa_cambio=Coalesce(
                'venta__radio_multiplicador_usado', 
                'venta__moneda__radioMultiplicador', 
                1,
                output_field=DecimalField()
            )
        ).annotate(
            tasa_final=Case(
                When(tasa_cambio=0, then=Value(1)),
                default=F('tasa_cambio'),
                output_field=DecimalField()
            )
        ).annotate(
            monto_usd=ExpressionWrapper(
                F('precio_total') / F('tasa_final'),
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )
        
        cats = detalles_qs.values('producto__codigo_categoria__nombre_categoria').annotate(
            total=Sum('monto_usd')
        ).order_by('-total')

        cat_labels = [c['producto__codigo_categoria__nombre_categoria'] for c in cats]
        cat_data = [float(c['total']) for c in cats]

        # Top Ranking Libros (Físico / Preventa) - considerar IDs y nombre de categoría
        libros_filter = DJANGO_Q(producto__codigo_categoria__in=[1, 15]) | DJANGO_Q(
            producto__codigo_categoria__nombre_categoria__icontains="fisico"
        ) | DJANGO_Q(producto__codigo_categoria__nombre_categoria__icontains="físico") | DJANGO_Q(
            producto__codigo_categoria__nombre_categoria__icontains="preventa"
        )
        top_books_qs = (
            DetalleVenta.objects.filter(
                libros_filter,
                venta__in=ventas_qs,
                producto__codigo_negocio__nombre_negocio__icontains="editorial"
            )
            .values("producto__nombre_producto")
            .annotate(total_qty=Sum("cantidad"))
            .order_by("-total_qty")[:10]
        )
        
        books_labels = [b['producto__nombre_producto'] for b in top_books_qs]
        books_data = [int(b['total_qty']) for b in top_books_qs]
        
        # Top Ranking Cursos (5=Online, 6=Virtual)
        top_courses_qs = DetalleVenta.objects.filter(
            venta__in=ventas_qs,
            producto__codigo_categoria__in=[5, 6]
        ).values('producto__nombre_producto').annotate(
            total_qty=Sum('cantidad')
        ).order_by('-total_qty')[:10]
        
        courses_labels = [b['producto__nombre_producto'] for b in top_courses_qs]
        courses_data = [int(b['total_qty']) for b in top_courses_qs]
        
        # Ranking Vendedores (cantidad de ventas + montos de cuotas en USD)
        try:
            vendors_qs = ventas_qs.values(
                'usuario_id', 'usuario__username', 'usuario__first_name', 'usuario__last_name'
            ).annotate(
                total_ventas=Count('id'),
                total_monto=Sum('monto_usd')
            ).order_by('-total_ventas', '-total_monto')[:20]

            # Deuda/pagado por vendedor desde cuotas (mismo universo del dashboard).
            # Deuda: cuotas estado=2 (Pendiente)
            # Pagado: cuotas estado=1 (Pagado)
            vendor_cuotas_qs = cuotas_qs.values('venta__usuario_id').annotate(
                deuda_usd=Sum(
                    Case(
                        When(estado=2, then=F("monto_usd")),
                        default=Value(Decimal("0.00")),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
                monto_pagado_usd=Sum(
                    Case(
                        When(estado=1, then=F("monto_usd")),
                        default=Value(Decimal("0.00")),
                        output_field=DecimalField(max_digits=12, decimal_places=2),
                    )
                ),
            )
            vendor_cuotas_map = {
                int(row["venta__usuario_id"]): {
                    "deuda_usd": float(row.get("deuda_usd") or 0),
                    "monto_pagado_usd": float(row.get("monto_pagado_usd") or 0),
                }
                for row in vendor_cuotas_qs
                if row.get("venta__usuario_id") is not None
            }

            vendors_data = []
            for v in vendors_qs:
                name = f"{v['usuario__first_name']} {v['usuario__last_name']}".strip() or v['usuario__username']
                cuota_stats = vendor_cuotas_map.get(int(v['usuario_id']), {"deuda_usd": 0.0, "monto_pagado_usd": 0.0})
                total_cuotas_usd = float(cuota_stats["deuda_usd"] or 0) + float(cuota_stats["monto_pagado_usd"] or 0)
                vendors_data.append({
                    "user_id": v['usuario_id'],
                    "username": v['usuario__username'],
                    "name": name,
                    "count": v['total_ventas'],
                    "deuda_usd": cuota_stats["deuda_usd"],
                    "monto_pagado_usd": cuota_stats["monto_pagado_usd"],
                    "monto_total_usd": total_cuotas_usd,
                    # Compatibilidad con frontend previo.
                    "amount": total_cuotas_usd,
                })
        except Exception:
            vendors_data = []

        # Ventas por país (basado en país del cliente; fallback a venta.pais)
        try:
            country_qs = ventas_qs.values(
                "cliente__pais__nombre", "pais__nombre"
            ).annotate(
                total_ventas=Count("id"),
                total_monto=Sum("monto_usd")
            )

            country_bucket = {}
            for row in country_qs:
                country_name = (row.get("cliente__pais__nombre") or row.get("pais__nombre") or "").strip() or "Sin país"
                bucket = country_bucket.setdefault(country_name, {"count": 0, "amount": 0.0})
                bucket["count"] += int(row.get("total_ventas") or 0)
                bucket["amount"] += float(row.get("total_monto") or 0)

            top_country = sorted(
                country_bucket.items(),
                key=lambda x: (-x[1]["count"], -x[1]["amount"], x[0])
            )[:12]

            country_labels = [k for k, _ in top_country]
            country_data = [v["count"] for _, v in top_country]
        except Exception:
            country_labels = []
            country_data = []

        # Ventas por medio
        try:
            medium_qs = ventas_qs.values("medio").annotate(
                total_ventas=Count("id")
            )

            medium_bucket = {}
            for row in medium_qs:
                label = _medio_label(row.get("medio"))
                medium_bucket[label] = medium_bucket.get(label, 0) + int(row.get("total_ventas") or 0)

            top_medium = sorted(
                medium_bucket.items(),
                key=lambda x: (-x[1], x[0])
            )[:10]

            medium_labels = [k for k, _ in top_medium]
            medium_data = [v for _, v in top_medium]
        except Exception:
            medium_labels = []
            medium_data = []

        ventas_chart_labels = [m["month"] for m in monthly]
        ventas_chart_data = [float(m["total"]) for m in monthly]
        
        if not ventas_chart_data and ventas_resumen.get("monto"):
            ventas_chart_labels = ["Actual"]
            ventas_chart_data = [float(ventas_resumen.get("monto") or 0)]
        
        cache.set(cache_key, {
            "ventas_resumen": ventas_resumen,
            "ventas_por_estado": ventas_por_estado,
            "monthly": monthly,
            "cuotas_por_estado": cuotas_por_estado,
            "ventas_chart_labels": ventas_chart_labels,
            "ventas_chart_data": ventas_chart_data,
            "ventas_estado_chart_data": ventas_estado_chart_data,
            "cuotas_estado_chart_data": cuotas_estado_chart_data,
            "cat_labels": cat_labels,
            "cat_data": cat_data,
            "books_labels": books_labels,
            "books_data": books_data,
            "courses_labels": courses_labels,
            "courses_data": courses_data,
            "vendors_data": vendors_data,
            "country_labels": country_labels,
            "country_data": country_data,
            "medium_labels": medium_labels,
            "medium_data": medium_data,
        }, 300)

    # ---- Meta KPIs (mes actual + mes anterior) ---- computed fresh, no caché
    # Admin-only global data; computed outside the cached block so it stays current.
    meta_kpis = None
    meta_kpis_prev = None
    show_meta_kpis = bool(is_admin or (settings.DEBUG and _is_local_request(request)))
    if show_meta_kpis:
        try:
            today = datetime.date.today()
            # Current month: 1st → today
            _cm_from = datetime.date(today.year, today.month, 1)
            _cm_to = today
            meta_kpis = _paid_media_kpis(_cm_from, _cm_to)
            # Previous month
            _prev_last = _cm_from - datetime.timedelta(days=1)
            _pm_from = datetime.date(_prev_last.year, _prev_last.month, 1)
            _pm_to = _prev_last
            meta_kpis_prev = _paid_media_kpis(_pm_from, _pm_to)
        except Exception:
            meta_kpis = None
            meta_kpis_prev = None

    # API Carga Asíncrona
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        # v12: incluye detalle_venta_monto_usd para prorrateo correcto de packs en frontend
        list_cache_key = f"dash_list_v12_pending_paid_{request.user.id}_{is_admin}"
        cached_lists = cache.get(list_cache_key)
        
        if cached_lists:
            return JsonResponse(cached_lists, encoder=DjangoJSONEncoder)

        # Para Vendedores dinámico en JS, necesitamos saber quien vendió cada venta.
        # Ya tenemos ventas_detalle, pero falta el nombre del vendedor.
        # Agregamos username a ventas_detalle query arriba? 
        # Mejor re-hacemos la query de ventas_detalle para incluir usuario info.
        
        ventas_detalle = list(
            ventas_qs.values(
                "id", "folio_venta", "monto_usd", "estado", "fecha_evento",
                "usuario_id", "usuario__username", "usuario__first_name", "usuario__last_name",
                "cliente__pais__nombre", "pais__nombre", "medio"
            )
        )
        cuotas_por_venta_qs = cuotas_qs.values("venta_id").annotate(
            cuota_pendiente_usd=Sum(
                Case(
                    When(estado=2, then=F("monto_usd")),
                    default=Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            cuota_pagada_usd=Sum(
                Case(
                    When(estado=1, then=F("monto_usd")),
                    default=Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
        )
        cuotas_por_venta_map = {
            int(row["venta_id"]): {
                "cuota_pendiente_usd": float(row.get("cuota_pendiente_usd") or 0),
                "cuota_pagada_usd": float(row.get("cuota_pagada_usd") or 0),
            }
            for row in cuotas_por_venta_qs
            if row.get("venta_id") is not None
        }
        for v in ventas_detalle:
            v["monto_total"] = v.pop("monto_usd")
            # Mantener compatibilidad con el frontend actual (usa v.fecha_venta)
            v["fecha_venta"] = v.pop("fecha_evento", None)
            v["vendedor"] = f"{v['usuario__first_name']} {v['usuario__last_name']}".strip() or v['usuario__username']
            v["pais_cliente"] = (v.pop("cliente__pais__nombre", None) or v.pop("pais__nombre", None) or "").strip() or "Sin país"
            v["medio_label"] = _medio_label(v.get("medio"))
            cuotas_venta = cuotas_por_venta_map.get(int(v["id"]), {"cuota_pendiente_usd": 0.0, "cuota_pagada_usd": 0.0})
            v["cuota_pendiente_usd"] = cuotas_venta["cuota_pendiente_usd"]
            v["cuota_pagada_usd"] = cuotas_venta["cuota_pagada_usd"]

        cuotas_detalle = list(
            cuotas_qs.values(
                "id", "venta__folio_venta", "monto_usd", "estado", "fecha_vencimiento",
            )
        )
        for c in cuotas_detalle:
            c["monto_total"] = c.pop("monto_usd")
        
        # Detalles con categoría para filtrar en JS
        # (Reutilizamos la query de detalles_qs pero sin agrupar, filtrando por ID venta)
        # Nota: detalles_qs ya depende de ventas_qs, asi que es consistente.
        # Pero ventas_qs es QuerySet, asi que podemos reconstruir detalles query rapido:
        
        detalles_export = DetalleVenta.objects.filter(venta__in=ventas_qs).select_related(
            'venta', 'venta__moneda', 'producto__codigo_categoria', 'producto__codigo_negocio'
        ).annotate(
            tasa_cambio=Coalesce('venta__radio_multiplicador_usado', 'venta__moneda__radioMultiplicador', 1, output_field=DecimalField())
        ).annotate(
            tasa_final=Case(When(tasa_cambio=0, then=Value(1)), default=F('tasa_cambio'), output_field=DecimalField())
        ).annotate(
            monto_usd=ExpressionWrapper(F('precio_total') / F('tasa_final'), output_field=DecimalField(max_digits=12, decimal_places=2))
        ).values(
            "venta_id", 
            "producto__codigo_producto",
            "producto__codigo_categoria__nombre_categoria", 
            "producto__codigo_categoria__nombre_categoria", 
            "monto_usd",
            "producto__codigo_categoria", # Para filtrar ID en JS
            "producto__codigo_negocio",
            "producto__codigo_negocio__nombre_negocio",
            "producto__nombre_producto",
            "producto__imagen_producto",
            "cantidad"
        )

        detalles_export = list(detalles_export)
        for row in detalles_export:
            row["producto_imagen_url"] = _resolve_media_url(row.pop("producto__imagen_producto", None))

        libros_en_pack_export = (
            LibroEnPack.objects.filter(detalle_venta__venta__in=ventas_qs)
            .select_related(
                "detalle_venta",
                "detalle_venta__venta",
                "detalle_venta__venta__moneda",
                "detalle_venta__producto",
                "libro",
            )
            .annotate(
                tasa_cambio=Coalesce(
                    "detalle_venta__venta__radio_multiplicador_usado",
                    "detalle_venta__venta__moneda__radioMultiplicador",
                    1,
                    output_field=DecimalField(),
                )
            )
            .annotate(
                tasa_final=Case(
                    When(tasa_cambio=0, then=Value(1)),
                    default=F("tasa_cambio"),
                    output_field=DecimalField(),
                )
            )
            .annotate(
                precio_unitario_usd=ExpressionWrapper(
                    F("precio_unitario") / F("tasa_final"),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
                detalle_venta_monto_usd=ExpressionWrapper(
                    F("detalle_venta__precio_total") / F("tasa_final"),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
                venta_fecha=Coalesce(
                    "detalle_venta__venta__fecha_venta",
                    "detalle_venta__venta__fecha_registro",
                    output_field=DateTimeField(),
                ),
                venta_id=F("detalle_venta__venta_id"),
                folio_venta=F("detalle_venta__venta__folio_venta"),
                pack_producto_id=F("detalle_venta__producto__codigo_producto"),
                pack_producto_nombre=F("detalle_venta__producto__nombre_producto"),
                pack_imagen_path=F("detalle_venta__producto__imagen_producto"),
                libro_nombre=F("libro__nombre_producto"),
                libro_imagen_path=F("libro__imagen_producto"),
            )
            .values(
                "detalle_venta_id",
                "venta_id",
                "venta_fecha",
                "folio_venta",
                "pack_producto_id",
                "pack_producto_nombre",
                "pack_imagen_path",
                "libro_id",
                "libro_nombre",
                "libro_imagen_path",
                "cantidad",
                "precio_unitario_usd",
                "detalle_venta_monto_usd",
            )
        )

        libros_en_pack_export = list(libros_en_pack_export)
        for row in libros_en_pack_export:
            row["pack_imagen_url"] = _resolve_media_url(row.pop("pack_imagen_path", None))
            row["libro_imagen_url"] = _resolve_media_url(row.pop("libro_imagen_path", None))

        response_data = {
            "ventas_detalle": ventas_detalle,
            "cuotas_detalle": cuotas_detalle,
            "ventas_estado_detalle": list(ventas_scope.values("estado")),
            "cuotas_estado_detalle": list(cuotas_scope.values("estado")),
            "detalles_categoria": detalles_export, 
            "libros_en_pack_detalle": libros_en_pack_export,
        }
        cache.set(list_cache_key, response_data, 300)
        return JsonResponse(response_data, encoder=DjangoJSONEncoder)
    
    # Perfil
    perfil = None
    try:
        # PerfilUsuario usa OneToOne con user, pero en managed=False puede fallar si no hay user instance real
        # Si estamos debugueando, omitimos perfil o lo buscamos manual
        if request.user.is_authenticated:
            perfil = PerfilUsuario.objects.get(user=request.user)
        else:
            # Fake perfil lookup if needed, or pass None
            pass
    except PerfilUsuario.DoesNotExist:
        pass

    username_str = "Visitante (Debug)"
    if request.user.is_authenticated:
        username_str = request.user.get_full_name() or request.user.username
    else:
        # Intentar simular nombre si quisieramos, pero 'Visitante' basta
        pass

    context = {
        "username": username_str,
        "perfil": perfil,
        "ventas_resumen": {
            "total": ventas_resumen.get("total") or 0,
            "monto": ventas_resumen.get("monto") or Decimal("0.00"),
        },
        "is_admin": is_admin,
        "ventas_por_estado": list(ventas_por_estado),
        "monthly": list(monthly),
        "cuotas_por_estado": list(cuotas_por_estado),
        "ventas_detalle_json": "[]", 
        "cuotas_detalle_json": "[]",
        "chart_ventas_labels": json.dumps(ventas_chart_labels, cls=DjangoJSONEncoder),
        "chart_ventas_data": json.dumps(ventas_chart_data, cls=DjangoJSONEncoder),
        "chart_ventas_estado": json.dumps(ventas_estado_chart_data, cls=DjangoJSONEncoder),
        "chart_ventas_estado": json.dumps(ventas_estado_chart_data, cls=DjangoJSONEncoder),
        "chart_cuotas_estado": json.dumps(cuotas_estado_chart_data, cls=DjangoJSONEncoder),
        "chart_cat_labels": json.dumps(cat_labels, cls=DjangoJSONEncoder),
        "chart_cat_labels": json.dumps(cat_labels, cls=DjangoJSONEncoder),
        "chart_cat_data": json.dumps(cat_data, cls=DjangoJSONEncoder),
        "chart_books_labels": json.dumps(books_labels, cls=DjangoJSONEncoder),
        "chart_books_data": json.dumps(books_data, cls=DjangoJSONEncoder),
        "chart_courses_labels": json.dumps(courses_labels, cls=DjangoJSONEncoder),
        "chart_courses_data": json.dumps(courses_data, cls=DjangoJSONEncoder),
        "chart_country_labels": json.dumps(country_labels, cls=DjangoJSONEncoder),
        "chart_country_data": json.dumps(country_data, cls=DjangoJSONEncoder),
        "chart_medium_labels": json.dumps(medium_labels, cls=DjangoJSONEncoder),
        "chart_medium_data": json.dumps(medium_data, cls=DjangoJSONEncoder),
        "vendors_data": json.dumps(vendors_data, cls=DjangoJSONEncoder),
        # Meta KPIs — paid-media funnel (mes actual); admin-only or DEBUG
        "show_meta_kpis": show_meta_kpis,
        "meta_kpis": meta_kpis,
        "meta_kpis_prev": meta_kpis_prev,
        # Variables extra para el template satélite
        "MAIN_APP_URL": getattr(settings, 'MAIN_APP_URL', ''),
    }
    return render(request, "home.html", context)


def _export_ventas(queryset, fmt: str = "csv"):
    # (Misma implementación de exportación)
    def _naive(dt):
        if not dt: return ""
        if hasattr(dt, "tzinfo") and dt.tzinfo: return dt.replace(tzinfo=None)
        return dt

    def _fecha_export(v):
        return getattr(v, "fecha_evento", None) or getattr(v, "fecha_venta", None) or getattr(v, "fecha_registro", None)

    if fmt in ("xls", "xlsx") and openpyxl:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"
        ws.append(["Folio", "Monto (USD)", "Estado", "Fecha"])
        for v in queryset:
            ws.append([
                v.folio_venta, float(v.monto_usd or 0), v.estado,
                _naive(_fecha_export(v)),
            ])
        buffer = BytesIO()
        wb.save(buffer)
        resp = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = "attachment; filename=ventas.xlsx"
        return resp
        
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Folio", "Monto (USD)", "Estado", "Fecha"])
    for v in queryset:
        writer.writerow([v.folio_venta, v.monto_usd, v.estado, _fecha_export(v)])
    resp = HttpResponse(buffer.getvalue(), content_type="text/csv")
    resp["Content-Disposition"] = "attachment; filename=ventas.csv"
    return resp


def _export_cuotas(queryset, fmt: str = "csv"):
    def _naive(dt):
        if not dt: return ""
        if hasattr(dt, "tzinfo") and dt.tzinfo: return dt.replace(tzinfo=None)
        return dt

    if fmt in ("xls", "xlsx") and openpyxl:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuotas"
        ws.append(["Venta", "Cuota", "Monto (USD)", "Estado", "Vence"])
        for c in queryset:
            # c.venta ya se trajo con select_related
            ws.append([
                c.venta.folio_venta, c.numero_cuota, float(c.monto_usd or 0), c.estado,
                _naive(c.fecha_vencimiento) if hasattr(c, "fecha_vencimiento") else "",
            ])
        buffer = BytesIO()
        wb.save(buffer)
        resp = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = "attachment; filename=cuotas.xlsx"
        return resp

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Venta", "Cuota", "Monto (USD)", "Estado", "Vence"])
    for c in queryset:
        writer.writerow([c.venta.folio_venta, c.numero_cuota, c.monto_usd, c.estado, c.fecha_vencimiento])
    resp = HttpResponse(buffer.getvalue(), content_type="text/csv")
    resp["Content-Disposition"] = "attachment; filename=cuotas.csv"
    return resp


# ---------------------------------------------------------------------------
# Ads / Pauta — Meta Ads view
# ---------------------------------------------------------------------------

# Column order for the 21-column Pauta table (APV-3)
ADS_COLUMNS = [
    ("campaign_name",   "Nombre de la Campaña"),
    ("product",         "Producto"),
    ("paid_country",    "País Pagado"),
    ("country",         "País"),
    ("delivery",        "Entrega"),
    ("results",         "Resultados"),
    ("result_indicator","Indicador de resultados"),
    ("reach",           "Alcance"),
    ("impressions",     "Impresiones"),
    ("link_clicks",     "Link Clicks"),
    ("cost_per_result", "Costo por Resultados"),
    ("spend",           "Importe Gastado"),
    ("month",           "Mes"),
    ("start_date",      "Inicio"),
    ("end_date",        "Fin"),
    ("report_start",    "Inicio del Informe"),
    ("report_end",      "Fin del Informe"),
    ("campaign_id",     "ID de la Campaña"),
    ("account_id",      "ID de la Cuenta"),
    ("category",        "Categoría"),
    ("amount_usd",      "Monto Dólares"),
]

ADS_FIELD_NAMES = [col[0] for col in ADS_COLUMNS]


def _serialize_ads_row(obj, account_map=None, campaign_status_map=None) -> dict:
    """Serialize a MetaAds instance to a JSON-safe dict for the AJAX payload."""
    def _date_str(d):
        if d is None:
            return None
        if hasattr(d, "isoformat"):
            return d.isoformat()
        return str(d)

    def _dec(v):
        if v is None:
            return None
        return float(v)

    acct = (account_map or {}).get(obj.account_id) if obj.account_id else None

    # effective_status: use own value if present, else fall back to any API row
    # with the same campaign_id (covers Excel rows not yet backfilled).
    eff_status = obj.effective_status
    if not eff_status and obj.campaign_id and campaign_status_map:
        eff_status = campaign_status_map.get(obj.campaign_id)

    return {
        "campaign_name":    obj.campaign_name,
        "product":          obj.product,
        "paid_country":     obj.paid_country,
        "country":          obj.country,
        "delivery":         obj.delivery,
        "effective_status": eff_status,
        "account_name":     acct["name"] if acct else None,
        "account_status":   acct["account_status"] if acct else None,
        "results":          obj.results,
        "result_indicator": obj.result_indicator,
        "reach":            obj.reach,
        "impressions":      obj.impressions,
        "link_clicks":      obj.link_clicks,
        "cost_per_result":  _dec(obj.cost_per_result),
        "spend":            _dec(obj.spend),
        "month":            obj.month,
        "start_date":       _date_str(obj.start_date),
        "end_date":         _date_str(obj.end_date),
        "report_start":     _date_str(obj.report_start),
        "report_end":       _date_str(obj.report_end),
        "campaign_id":      obj.campaign_id,
        "account_id":       obj.account_id,
        "category":         obj.category,
        "amount_usd":       _dec(obj.amount_usd),
        "source":           obj.source,
    }


def _export_meta_ads(queryset):
    """
    Export a MetaAds queryset to an openpyxl XLSX file.
    Returns an HttpResponse with the file attached.
    NULL values are rendered as an em-dash in the exported file.
    """
    if openpyxl is None:
        return HttpResponse("openpyxl not installed", status=500)

    today = datetime.date.today().isoformat()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pauta Meta"

    # Header row
    ws.append([label for _, label in ADS_COLUMNS])

    em_dash = "—"

    for obj in queryset:
        row = []
        for field, _ in ADS_COLUMNS:
            val = getattr(obj, field, None)
            if val is None:
                row.append(em_dash)
            elif hasattr(val, "isoformat"):
                row.append(val.isoformat())
            else:
                row.append(val)
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f"attachment; filename=pauta_meta_{today}.xlsx"
    return resp


@ads_admin_required
def ads_dashboard(request):
    """
    Dashboard de Pauta — Meta Ads.
    - GET (browser): renders ads.html
    - GET with X-Requested-With: XMLHttpRequest: returns JSON payload (paginated rows + filter options)
    - GET with ?export=xlsx: downloads filtered queryset as XLSX
    """
    # --- Build base queryset with optional filters ---
    qs = MetaAds.objects.all()

    mes = request.GET.get("mes", "").strip()
    pais = request.GET.get("pais", "").strip()
    categoria = request.GET.get("categoria", "").strip()
    producto = request.GET.get("producto", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if mes:
        qs = qs.filter(month=mes)
    if pais:
        qs = qs.filter(country=pais)
    if categoria:
        qs = qs.filter(category=categoria)
    if producto:
        qs = qs.filter(product=producto)
    if date_from:
        qs = qs.filter(report_start__gte=date_from)
    if date_to:
        qs = qs.filter(report_start__lte=date_to)

    # --- Export ---
    export = request.GET.get("export", "").strip().lower()
    if export in ("xlsx", "xls"):
        return _export_meta_ads(qs)

    # --- ROAS por producto (JSON) ---
    if request.GET.get("roas", "").strip() == "1":
        today = datetime.date.today()
        try:
            r_from = datetime.date.fromisoformat(date_from) if date_from else datetime.date(today.year, 1, 1)
            r_to = datetime.date.fromisoformat(date_to) if date_to else today
        except ValueError:
            return JsonResponse({"error": "fecha inválida"}, status=400)
        return JsonResponse({
            "date_from": r_from.isoformat(),
            "date_to": r_to.isoformat(),
            "rows": _roas_por_producto(r_from, r_to),
        })

    # --- AJAX JSON payload ---
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        # Distinct filter options come from the FULL queryset (no active filters)
        full_qs = MetaAds.objects.all()
        filters = {
            "mes":       sorted(filter(None, full_qs.values_list("month", flat=True).distinct())),
            "pais":      sorted(filter(None, full_qs.values_list("country", flat=True).distinct())),
            "categoria": sorted(filter(None, full_qs.values_list("category", flat=True).distinct())),
            "producto":  sorted(filter(None, full_qs.values_list("product", flat=True).distinct())),
        }

        # --- Full-dump mode: ?page_size=0 (or ?all=1) returns ALL filtered ---
        # rows unpaginated, ordered by id for stable output.
        # Serialization contract (via _serialize_ads_row): DateField -> ISO
        # "YYYY-MM-DD" string, Decimal -> JSON *number* (explicit float()),
        # NULL -> JSON null. The DjangoJSONEncoder never sees raw Decimals.
        # account_map: account_id -> {name, account_status}
        _acct_map = {
            a.account_id: {"name": a.name, "account_status": a.account_status}
            for a in MetaAccount.objects.only("account_id", "name", "account_status")
        }
        # campaign_status_map: campaign_id -> effective_status (from API rows only)
        # Used to fill in effective_status for Excel rows that have no API data yet.
        _campaign_status_map = {
            row["campaign_id"]: row["effective_status"]
            for row in MetaAds.objects.filter(
                source="api", effective_status__isnull=False, campaign_id__isnull=False
            ).values("campaign_id", "effective_status")
        }

        if (
            request.GET.get("all", "").strip() == "1"
            or request.GET.get("page_size", "").strip() == "0"
        ):
            rows = [_serialize_ads_row(obj, _acct_map, _campaign_status_map) for obj in qs.order_by("id")]
            return JsonResponse(
                {
                    "rows": rows,
                    "total": len(rows),
                    "page": 1,
                    "page_size": len(rows) or 1,
                    "filters": filters,
                },
                encoder=DjangoJSONEncoder,
            )

        # --- Default paginated mode (unchanged behavior) ---
        try:
            page = max(1, int(request.GET.get("page", 1)))
        except (ValueError, TypeError):
            page = 1
        try:
            page_size = min(500, max(1, int(request.GET.get("page_size", 200))))
        except (ValueError, TypeError):
            page_size = 200

        total = qs.count()
        offset = (page - 1) * page_size
        page_qs = qs[offset: offset + page_size]

        rows = [_serialize_ads_row(obj, _acct_map, _campaign_status_map) for obj in page_qs]

        return JsonResponse(
            {
                "rows": rows,
                "total": total,
                "page": page,
                "page_size": page_size,
                "filters": filters,
            },
            encoder=DjangoJSONEncoder,
        )

    # --- HTML render ---
    total = qs.count()

    # Count unlinked campaigns + spend context for toolbar badge.
    # Scoped to 2026 so the pill count matches the pending list in /ads/vincular/.
    try:
        _prog = _linking_progress(date_from=LINKING_SCOPE_FROM)
        unlinked_count     = _prog["unlinked_campaigns"]
        pct_spend_unlinked = _prog["pct_spend_unlinked"]
    except Exception:
        unlinked_count     = 0
        pct_spend_unlinked = 0

    # --- Freshness badge (item A1) ---
    try:
        last_synced = (
            MetaAds.objects
            .filter(source="api", synced_at__isnull=False)
            .order_by("-synced_at")
            .values_list("synced_at", flat=True)
            .first()
        )
    except Exception:
        last_synced = None

    freshness_label = None
    freshness_level = None  # "green" | "amber" | "red"
    freshness_title = None
    if last_synced is not None:
        now = timezone.now()
        delta = now - last_synced
        hours = delta.total_seconds() / 3600
        minutes = int(delta.total_seconds() / 60)
        if hours < 1:
            freshness_label = f"Actualizado hace {minutes} min"
        elif hours < 24:
            h = int(hours)
            freshness_label = f"Actualizado hace {h} h"
        else:
            d = int(hours / 24)
            freshness_label = f"Actualizado hace {d} día{'s' if d != 1 else ''}"
        if hours < 26:
            freshness_level = "green"
        elif hours < 50:
            freshness_level = "amber"
        else:
            freshness_level = "red"
        # Show exact local datetime in title tooltip
        local_dt = timezone.localtime(last_synced)
        freshness_title = local_dt.strftime("%d/%m/%Y %H:%M")

    context = {
        "total":             total,
        "mes":               mes,
        "pais":              pais,
        "categoria":         categoria,
        "producto":          producto,
        "unlinked_count":    unlinked_count,
        "pct_spend_unlinked": pct_spend_unlinked,
        "MAIN_APP_URL":      getattr(settings, "MAIN_APP_URL", ""),
        "freshness_label":   freshness_label,
        "freshness_level":   freshness_level,
        "freshness_title":   freshness_title,
    }
    return render(request, "ads.html", context)


# ---------------------------------------------------------------------------
# Ads / Cuentas Meta — account portfolio coverage view
# ---------------------------------------------------------------------------

# Shared with sync_meta_ads — import the constant to avoid drift.
# We import lazily so the view module doesn't require the META_ACCESS_TOKEN at
# load time (it may be absent in local dev / Vercel preview builds).
def _meta_api_version() -> str:
    try:
        from core.management.commands.sync_meta_ads import META_API_VERSION
        return META_API_VERSION
    except ImportError:
        return "v25.0"


# Meta account_status → human-readable Spanish label
_ACCOUNT_STATUS_LABELS = {
    1:   ("Activa",        "success"),
    2:   ("Deshabilitada", "secondary"),
    3:   ("Sin liquidar",  "warning"),
    7:   ("Archivada",     "secondary"),
    9:   ("En revisión",   "info"),
    100: ("Cerrada",       "dark"),
    101: ("Cerrada",       "dark"),
}

# 13-digit prefix match: the Excel era stored floats that corrupted the last
# 1-2 digits of 15-digit account IDs.  A shared 13-char prefix is always safe.
_ACCOUNT_PREFIX_LEN = 13


def _account_prefix(account_id: str) -> str:
    return str(account_id or "")[:_ACCOUNT_PREFIX_LEN]


def _build_spend_by_prefix(six_months_ago: datetime.date) -> dict[str, dict]:
    """
    Aggregate spend from tb_meta_ads (any source) per account, grouped by
    13-digit prefix of account_id. Returns {prefix: {current_month, prev_month}}.
    """
    today = datetime.date.today()
    current_month_start = datetime.date(today.year, today.month, 1)
    if today.month == 1:
        prev_month_start = datetime.date(today.year - 1, 12, 1)
        prev_month_end = datetime.date(today.year, 1, 1) - datetime.timedelta(days=1)
    else:
        prev_month_start = datetime.date(today.year, today.month - 1, 1)
        prev_month_end = current_month_start - datetime.timedelta(days=1)

    result: dict[str, dict] = {}

    rows = (
        MetaAds.objects
        .filter(report_start__gte=six_months_ago)
        .exclude(account_id__isnull=True)
        .values("account_id", "report_start", "spend")
    )
    for row in rows:
        prefix = _account_prefix(row["account_id"] or "")
        if not prefix:
            continue
        entry = result.setdefault(prefix, {"current_month": Decimal("0"), "prev_month": Decimal("0")})
        rs = row.get("report_start")
        sp = row.get("spend") or Decimal("0")
        if rs is None:
            continue
        if rs >= current_month_start:
            entry["current_month"] += sp
        elif prev_month_start <= rs <= prev_month_end:
            entry["prev_month"] += sp

    return result


def _call_graph_me_adaccounts(token: str) -> tuple[list[dict], str | None]:
    """
    Call GET /me/adaccounts with a 15-second timeout.
    Returns (accounts_list, error_message_or_None).
    """
    version = _meta_api_version()
    url = (
        f"https://graph.facebook.com/{version}/me/adaccounts"
        f"?fields=id,name,account_status,currency&limit=100"
        f"&access_token={urllib.parse.quote(token, safe='')}"
    )
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "goberna-dashboard/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        accounts = data.get("data", [])
        # Follow paging if needed (unlikely for <30 accounts)
        next_url = (data.get("paging") or {}).get("next")
        while next_url:
            req2 = urllib.request.Request(next_url, headers={"User-Agent": "goberna-dashboard/1.0"})
            with urllib.request.urlopen(req2, timeout=15) as resp2:
                page = json.loads(resp2.read().decode("utf-8"))
            accounts.extend(page.get("data", []))
            next_url = (page.get("paging") or {}).get("next")
        return accounts, None
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode("utf-8", errors="replace")
            err_obj = json.loads(body).get("error", {}) or {}
            msg = err_obj.get("message") or body[:300]
        except Exception:
            msg = str(e)
        return [], f"Error de API ({e.code}): {msg}"
    except Exception as e:
        return [], f"Error de red: {str(e)[:300]}"


def _upsert_accounts_from_api(accounts: list[dict]) -> tuple[int, int]:
    """Upsert list of dicts from Graph API into MetaAccount. Returns (inserted, updated)."""
    now = timezone.now()
    inserted = updated = 0
    existing_ids = set(MetaAccount.objects.values_list("account_id", flat=True))
    for acct in accounts:
        act_id = acct.get("id", "")
        acct_num = act_id[4:] if act_id.startswith("act_") else act_id
        name = acct.get("name") or act_id
        currency = acct.get("currency")
        status = acct.get("account_status")
        if acct_num not in existing_ids:
            MetaAccount.objects.create(
                account_id=acct_num,
                name=name,
                currency=currency,
                account_status=status,
                first_seen=now,
                last_seen=now,
            )
            inserted += 1
        else:
            MetaAccount.objects.filter(account_id=acct_num).update(
                name=name,
                currency=currency,
                account_status=status,
                last_seen=now,
            )
            updated += 1
    return inserted, updated


@ads_admin_required
def ads_accounts(request):
    """
    /ads/cuentas/ — Meta ad account portfolio coverage.

    ?check=1 → server-side verification: calls Graph API me/adaccounts,
               upserts tb_meta_accounts, redirects back with a banner.
    """
    token = os.getenv("META_ACCESS_TOKEN", "")
    token_present = bool(token)
    api_version = _meta_api_version()

    # ---- "Verificar conexión" action ----
    check = request.GET.get("check", "").strip() == "1"
    banner_info = None     # (message_str, level)  level: success | warning | danger

    if check:
        if not token_present:
            banner_info = (
                "Token META_ACCESS_TOKEN no configurado. No es posible verificar la conexión.",
                "warning",
            )
        else:
            api_accounts, api_error = _call_graph_me_adaccounts(token)
            if api_error:
                banner_info = (f"Error al contactar Meta API: {api_error}", "danger")
            else:
                prev_count = MetaAccount.objects.count()
                ins, upd = _upsert_accounts_from_api(api_accounts)
                new_count = MetaAccount.objects.count()
                new_accounts = new_count - prev_count
                banner_info = (
                    f"{len(api_accounts)} cuentas visibles ({new_accounts} nuevas, {upd} actualizadas).",
                    "success",
                )
        return redirect(f"/ads/cuentas/?_banner={'|'.join(banner_info)}")

    # Recover banner from redirect query param
    _banner_param = request.GET.get("_banner", "")
    if _banner_param and "|" in _banner_param:
        idx = _banner_param.index("|")
        banner_info = (_banner_param[:idx], _banner_param[idx + 1:])

    # ---- Data ----
    six_months_ago = datetime.date.today() - datetime.timedelta(days=183)

    # Last sync timestamp
    last_sync = (
        MetaAds.objects
        .filter(source="api", synced_at__isnull=False)
        .order_by("-synced_at")
        .values_list("synced_at", flat=True)
        .first()
    )

    # Linked accounts from tb_meta_accounts
    linked_accounts = list(MetaAccount.objects.all().order_by("name"))

    # Spend aggregated by 13-digit prefix
    spend_by_prefix = _build_spend_by_prefix(six_months_ago)

    # Enrich each linked account with spend data + status label
    linked_prefixes: set[str] = set()
    for acct in linked_accounts:
        prefix = _account_prefix(acct.account_id)
        linked_prefixes.add(prefix)
        sp = spend_by_prefix.get(prefix, {})
        acct.spend_current = sp.get("current_month", Decimal("0"))
        acct.spend_prev = sp.get("prev_month", Decimal("0"))
        label, badge = _ACCOUNT_STATUS_LABELS.get(
            acct.account_status, ("Otro", "secondary")
        )
        acct.status_label = label
        acct.status_badge = badge

    # ---- Unlinked portfolios ----
    # DISTINCT account_id from tb_meta_ads with recent data whose 13-digit
    # prefix does NOT match any known linked account.
    recent_ads_ids = (
        MetaAds.objects
        .filter(report_start__gte=six_months_ago)
        .exclude(account_id__isnull=True)
        .values_list("account_id", flat=True)
        .distinct()
    )

    unlinked_portfolios = []
    seen_prefixes: set[str] = set()
    for raw_id in recent_ads_ids:
        prefix = _account_prefix(raw_id)
        if not prefix or prefix in linked_prefixes or prefix in seen_prefixes:
            continue
        seen_prefixes.add(prefix)
        # Aggregate campaigns and spend for this prefix
        rows = (
            MetaAds.objects
            .filter(report_start__gte=six_months_ago, account_id__startswith=prefix)
            .values("campaign_name", "spend", "account_id")
        )
        campaign_names: list[str] = []
        total_spend = Decimal("0")
        campaign_ids: set[str] = set()
        for row in rows:
            sp = row.get("spend") or Decimal("0")
            total_spend += sp
            cn = row.get("campaign_name") or ""
            if cn and cn not in campaign_ids:
                campaign_names.append(cn)
                campaign_ids.add(cn)

        unlinked_portfolios.append({
            "account_id": raw_id,
            "prefix": prefix,
            "num_campaigns": len(campaign_ids),
            "total_spend": total_spend,
            "sample_campaigns": campaign_names[:5],
        })

    # Sort unlinked by spend desc
    unlinked_portfolios.sort(key=lambda x: -x["total_spend"])

    context = {
        "token_present": token_present,
        "api_version": api_version,
        "last_sync": last_sync,
        "linked_accounts": linked_accounts,
        "unlinked_portfolios": unlinked_portfolios,
        "banner_info": banner_info,
        "MAIN_APP_URL": getattr(settings, "MAIN_APP_URL", ""),
    }
    return render(request, "ads_accounts.html", context)


# ---------------------------------------------------------------------------
# Ads / Vincular productos — campaign-to-product linking screen
# ---------------------------------------------------------------------------

def _get_all_products():
    """Return all Producto rows as a list of dicts for the datalist picker."""
    return list(
        Producto.objects.select_related("codigo_negocio")
        .order_by("sku_producto")
        .values("codigo_producto", "sku_producto", "nombre_producto",
                "codigo_negocio__nombre_negocio")
    )


def _negocio_to_ads_category(nombre_negocio: str | None) -> str | None:
    """Normalize negocio name to the canonical category string used in tb_meta_ads."""
    import unicodedata
    if not nombre_negocio:
        return None
    key = "".join(
        ch for ch in unicodedata.normalize("NFD", nombre_negocio)
        if unicodedata.category(ch) != "Mn"
    ).lower()
    _MAP = {
        "consultoria": "Consultoría",
        "escuela":     "Escuela",
        "editorial":   "Editorial",
        "lifestyle":   "LifeStyle",
    }
    return _MAP.get(key, nombre_negocio)


# ---------------------------------------------------------------------------
# Vincular — agrupación de campañas pendientes + sugerencias de producto
# ---------------------------------------------------------------------------

_VIN_BRACKET_RE    = re.compile(r"\[([^\]]*)\]")
_VIN_SEP_RE        = re.compile(r"[|,\-_/]+")
_VIN_YEAR_RE       = re.compile(r"\b(?:19|20)\d{2}\b")
_VIN_MONTH_FULL_RE = re.compile(
    r"\b(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto"
    r"|septiembre|setiembre|octubre|noviembre|diciembre)\b"
)
# Fragmentos de fecha: "08 ene", "01nov", "22oct" (incluso pegados a una palabra)
_VIN_DATE_FRAG_RE  = re.compile(
    r"\d{1,2}\s*(?:ene|feb|mar|abr|may|jun|jul|ago|sep|set|oct|nov|dic)(?![a-z])"
)
_VIN_MONTH_ABBR_RE = re.compile(
    r"\b(?:ene|feb|mar|abr|may|jun|jul|ago|sep|set|oct|nov|dic)\b"
)
_VIN_NUM_RE        = re.compile(r"\b\d+\b")
# Tokens de canal/variante que no identifican el producto
_VIN_CHANNEL_RE    = re.compile(
    r"\b(?:wsp|whatsapp|int|land|landing|rec|ventas|venta|conversion"
    r"|interaccion|segmentado|videos|general|pp|pt|cp|wp|form|only"
    r"|new|pixel|leads|clientep|cliente)\b"
)


def _vin_strip_accents_lower(s: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFD", s or "")
        if unicodedata.category(ch) != "Mn"
    ).lower()


def _vin_norm_simple(s: str) -> str:
    """Normalización suave: sin acentos, lowercase, separadores → espacio."""
    return " ".join(_VIN_SEP_RE.sub(" ", _vin_strip_accents_lower(s)).split())


def _campaign_group_key(name: str) -> tuple[str, list[str]]:
    """
    Normaliza un nombre de campaña a una clave base de agrupación.
    Devuelve (key, bracket_codes): los códigos entre corchetes se quitan de la
    clave pero se conservan para el motor de sugerencias (match por SKU).
    """
    raw = name or ""
    bracket_codes: list[str] = []
    for seg in _VIN_BRACKET_RE.findall(raw):
        for tok in re.split(r"[\s,;]+", seg):
            tok = tok.strip().upper()
            if tok:
                bracket_codes.append(tok)
    s = _VIN_BRACKET_RE.sub(" ", raw)
    s = _vin_strip_accents_lower(s)
    s = _VIN_SEP_RE.sub(" ", s)
    s = _VIN_YEAR_RE.sub(" ", s)
    s = _VIN_MONTH_FULL_RE.sub(" ", s)
    s = _VIN_DATE_FRAG_RE.sub(" ", s)
    s = _VIN_MONTH_ABBR_RE.sub(" ", s)
    s = _VIN_NUM_RE.sub(" ", s)
    s = _VIN_CHANNEL_RE.sub(" ", s)
    key = " ".join(s.split())
    if not key:
        key = _vin_norm_simple(raw) or raw
    return key, bracket_codes


def _group_pending_campaigns(unlinked: list[dict],
                             sku_index: dict | None = None) -> list[dict]:
    """
    Agrupa campañas pendientes por clave normalizada.
    Si el nombre de la campaña trae un código bracketed que existe como SKU en
    el catálogo, el grupo se particiona por ese SKU: ediciones/productos
    distintos ([DIPICOT016] vs [DIPICOT017] vs [PKCONTR001]) nunca se mezclan
    en un mismo bulk-link aunque su clave normalizada coincida.
    Cada grupo: title (nombre del miembro más reciente), members (orden gasto
    desc), n, total_spend, search_blob y _code_stats (códigos bracketed con
    frecuencia y gasto, para sugerencias). Orden: gasto total desc.
    """
    sku_index = sku_index or {}
    groups: dict[str, dict] = {}
    for camp in unlinked:
        key, codes = _campaign_group_key(camp.get("campaign_name") or "")
        # Partición por SKU de catálogo: campañas cuyo propio código resuelve
        # a un producto forman grupo aparte (clave key + sku).
        resolved_sku = next((c for c in codes if c in sku_index), None)
        group_key = f"{key}::sku::{resolved_sku}" if resolved_sku else key
        g = groups.get(group_key)
        if g is None:
            g = groups[group_key] = {
                "key": key,
                "members": [],
                "total_spend": 0.0,
                "_code_stats": {},
            }
        spend = float(camp.get("recent_spend") or 0)
        g["members"].append(camp)
        g["total_spend"] += spend
        for code in codes:
            st = g["_code_stats"].setdefault(code, {"n": 0, "spend": 0.0})
            st["n"] += 1
            st["spend"] += spend

    out: list[dict] = []
    for g in groups.values():
        g["members"].sort(key=lambda m: -(m.get("recent_spend") or 0))
        title_member = max(
            g["members"],
            key=lambda m: (m.get("last_seen") or datetime.date.min,
                           m.get("recent_spend") or 0),
        )
        g["title"] = title_member.get("campaign_name") or g["key"]
        g["n"] = len(g["members"])
        g["search_blob"] = " ".join(
            f"{m.get('campaign_name') or ''} {m.get('campaign_id') or ''}"
            for m in g["members"]
        ).lower()
        out.append(g)

    out.sort(key=lambda x: -x["total_spend"])
    return out


# Fuzzy: umbral mínimo y productos genéricos/placeholder del catálogo que
# nunca deben ofrecerse como sugerencia ("Consultoría" GEN09EA5D es genérico,
# "Libro Prueba" CODIGO es un producto de prueba).
_VIN_FUZZY_THRESHOLD = 0.75
_VIN_FUZZY_SKU_BLACKLIST = {"GEN09EA5D", "CODIGO"}


def _vin_shared_distinctive_token(a: str, b: str) -> bool:
    """True si ambos textos normalizados comparten al menos un token de 4+
    caracteres (evita sugerencias por similitud superficial de letras)."""
    ta = {t for t in a.split() if len(t) >= 4}
    tb = {t for t in b.split() if len(t) >= 4}
    return bool(ta & tb)


def _suggest_product_for_group(group: dict, sku_index: dict,
                               name_index: list) -> dict | None:
    """
    Sugerencia de producto para un grupo:
      (a) código bracketed que matchea Producto.sku_producto exacto
          (el más frecuente entre los miembros gana) → confianza "SKU";
      (b) fuzzy: clave del grupo vs nombre de producto normalizado,
          SequenceMatcher ratio >= 0.75 + al menos un token distintivo
          compartido, excluyendo productos placeholder → confianza "similar";
      (c) claves con libro/pack/libreria sin mejor match → producto "Librería"
          del catálogo si existe.
    """
    # (a) SKU exacto en corchetes — más frecuente primero, luego más gasto
    candidates = sorted(
        group.get("_code_stats", {}).items(),
        key=lambda kv: (-kv[1]["n"], -kv[1]["spend"], kv[0]),
    )
    for code, _st in candidates:
        p = sku_index.get(code)
        if p:
            return {
                "codigo": p["codigo_producto"],
                "nombre": p["nombre_producto"],
                "sku": p["sku_producto"],
                "confidence": "SKU",
            }

    # (b) fuzzy contra nombres de producto normalizados
    key = group["key"]
    best, best_ratio = None, 0.0
    sm = SequenceMatcher(None, "", key)  # seq2 fija (cacheada por difflib)
    for pn_norm, p in name_index:
        if not pn_norm:
            continue
        if (p["sku_producto"] or "").strip().upper() in _VIN_FUZZY_SKU_BLACKLIST:
            continue
        if not _vin_shared_distinctive_token(pn_norm, key):
            continue
        sm.set_seq1(pn_norm)
        if (sm.real_quick_ratio() < _VIN_FUZZY_THRESHOLD
                or sm.quick_ratio() < _VIN_FUZZY_THRESHOLD):
            continue
        ratio = sm.ratio()
        if ratio > best_ratio:
            best_ratio, best = ratio, p
    if best is not None and best_ratio >= _VIN_FUZZY_THRESHOLD:
        return {
            "codigo": best["codigo_producto"],
            "nombre": best["nombre_producto"],
            "sku": best["sku_producto"],
            "confidence": "similar",
        }

    # (c) libro/pack/libreria → producto "Librería" del catálogo, si existe
    if any(t in key for t in ("libro", "pack", "libreria")):
        p = next(
            (p for pn_norm, p in name_index
             if "libreria" in pn_norm
             and (p["sku_producto"] or "").strip().upper()
             not in _VIN_FUZZY_SKU_BLACKLIST),
            None,
        )
        if p:
            return {
                "codigo": p["codigo_producto"],
                "nombre": p["nombre_producto"],
                "sku": p["sku_producto"],
                "confidence": "similar",
            }
    return None


def _linking_progress(date_from=None) -> dict:
    """
    Compute campaign-linking progress stats.

    Args:
        date_from: optional datetime.date.  When given, all aggregates
                   (total/linked campaigns, spend totals/pcts, top-10)
                   consider ONLY MetaAds rows with report_start >= date_from.
                   Pass LINKING_SCOPE_FROM (2026-01-01) from callers that want
                   the 2026-scoped view.  Default None = global (all rows).

    Returns a dict with:
      total_campaigns   – distinct campaign_ids in the (optionally filtered) rows
      linked_campaigns  – campaign_ids with a map entry (codigo_producto NOT NULL)
      unlinked_campaigns
      pct_campaigns     – int 0-100
      spend_total       – Decimal: sum of spend for rows in scope
      spend_linked      – Decimal: sum of spend for linked campaign_ids in scope
      spend_unlinked    – Decimal
      pct_spend         – int 0-100
      pct_spend_unlinked – int 0-100 (convenience for ads_dashboard pill)
      top10_unlinked_spend – Decimal: spend of the top-10 unlinked campaigns by spend
      top10_pct_extra    – int: what % of total spend would those 10 add to linked
    All monetary values are Decimal; percentages are Python ints (floor).
    """
    from decimal import Decimal
    from django.db.models import Sum as DSum

    zero = Decimal("0")

    # Linked campaign_ids (have a map entry with producto) — always global:
    # a campaign linked at any time counts as linked regardless of date scope.
    linked_ids = set(
        MetaCampaignMap.objects
        .exclude(codigo_producto__isnull=True)
        .values_list("campaign_id", flat=True)
    )

    # All distinct campaigns + their total spend (scoped if date_from given)
    base_qs = MetaAds.objects.exclude(campaign_id__isnull=True)
    if date_from is not None:
        base_qs = base_qs.filter(report_start__gte=date_from)
    all_camp_spend = (
        base_qs
        .values("campaign_id")
        .annotate(total_spend=DSum("spend"))
    )

    total_campaigns  = 0
    linked_campaigns = 0
    spend_total      = zero
    spend_linked     = zero

    # Collect unlinked campaigns with their spend for the top-10 hint
    unlinked_spends: list[Decimal] = []

    for row in all_camp_spend:
        cid = row["campaign_id"]
        sp  = row["total_spend"] or zero
        total_campaigns += 1
        spend_total += sp
        if cid in linked_ids:
            linked_campaigns += 1
            spend_linked += sp
        else:
            unlinked_spends.append(sp)

    unlinked_campaigns = total_campaigns - linked_campaigns
    spend_unlinked     = spend_total - spend_linked

    pct_campaigns      = int(linked_campaigns * 100 // total_campaigns) if total_campaigns else 0
    pct_spend          = int(spend_linked * 100 // spend_total)         if spend_total    else 0
    pct_spend_unlinked = 100 - pct_spend

    # Top-10 unlinked campaigns by spend
    unlinked_spends.sort(reverse=True)
    top10_unlinked_spend = sum(unlinked_spends[:10], zero)
    top10_pct_extra = int(top10_unlinked_spend * 100 // spend_total) if spend_total else 0

    return {
        "total_campaigns":      total_campaigns,
        "linked_campaigns":     linked_campaigns,
        "unlinked_campaigns":   unlinked_campaigns,
        "pct_campaigns":        pct_campaigns,
        "spend_total":          spend_total,
        "spend_linked":         spend_linked,
        "spend_unlinked":       spend_unlinked,
        "pct_spend":            pct_spend,
        "pct_spend_unlinked":   pct_spend_unlinked,
        "top10_unlinked_spend": top10_unlinked_spend,
        "top10_pct_extra":      top10_pct_extra,
    }


@ads_admin_required
def ads_vincular(request):
    """
    /ads/vincular/ — campaign-to-product linking screen.

    GET:  List unlinked campaigns (no map entry with codigo_producto) grouped
          by normalized name, ordered by group spend desc.
    POST (CSRF):
      - bulk (campaign_ids repetidos + codigo_producto/codigo_sugerido):
        upsert map entries linked_by='manual' + update tb_meta_ads, atómico.
      - single (campaign_id): flujo original, intacto (re-vincular incluido).
    """
    banner = None  # (message, level)

    if request.method == "POST" and (
        request.POST.get("bulk") == "1" or request.POST.getlist("campaign_ids")
    ):
        # ---- Bulk link: N campañas de un grupo → un producto ----
        bulk_ids = [c.strip() for c in request.POST.getlist("campaign_ids") if c.strip()]
        # Dedupe conservando el orden: ids repetidos inflarían el contador
        # del banner y duplicarían escrituras.
        bulk_ids = list(dict.fromkeys(bulk_ids))
        codigo_str = (request.POST.get("codigo_sugerido")
                      or request.POST.get("codigo_producto") or "").strip()

        if not bulk_ids:
            banner = ("Seleccioná al menos una campaña del grupo.", "danger")
        elif len(bulk_ids) > 200:
            banner = ("Demasiadas campañas en una sola vinculación (máximo 200).", "danger")
        elif not codigo_str:
            banner = ("Faltan datos: producto.", "danger")
        else:
            try:
                codigo_producto = int(codigo_str)
            except ValueError:
                codigo_producto = None
                banner = ("Código de producto inválido.", "danger")

            if codigo_producto is not None:
                prod = (
                    Producto.objects.select_related("codigo_negocio")
                    .filter(codigo_producto=codigo_producto)
                    .first()
                )
                if prod is None:
                    banner = ("El producto seleccionado no existe en el catálogo.", "danger")
                else:
                    # Validar que todos los ids existan en tb_meta_ads antes
                    # de escribir: ids forjados o de páginas viejas crearían
                    # filas huérfanas (invisibles en la UI) en la tabla
                    # compartida de producción.
                    valid_ids = set(
                        MetaAds.objects
                        .filter(campaign_id__in=bulk_ids)
                        .values_list("campaign_id", flat=True)
                    )
                    unknown = [c for c in bulk_ids if c not in valid_ids]
                    if unknown:
                        banner = (
                            "Algunas campañas seleccionadas no existen en los "
                            "datos de Meta. Recargá la página e intentá de nuevo.",
                            "danger",
                        )
                    else:
                        product_name = prod.nombre_producto
                        try:
                            negocio = prod.codigo_negocio.nombre_negocio
                        except Exception:
                            negocio = None
                        category = _negocio_to_ads_category(negocio)
                        try:
                            from django.db import transaction
                            from django.utils import timezone as tz
                            with transaction.atomic():
                                for cid in bulk_ids:
                                    MetaCampaignMap.objects.update_or_create(
                                        campaign_id=cid,
                                        defaults={
                                            "codigo_producto": codigo_producto,
                                            "product_name": product_name or None,
                                            "category": category or None,
                                            "linked_by": "manual",
                                            "linked_at": tz.now(),
                                        },
                                    )
                                MetaAds.objects.filter(campaign_id__in=bulk_ids).update(
                                    product=product_name or None,
                                    category=category or None,
                                )
                            _prog_post = _linking_progress(date_from=LINKING_SCOPE_FROM)
                            n = len(bulk_ids)
                            plural = "s" if n != 1 else ""
                            banner = (
                                f"✓ {n} campaña{plural} vinculada{plural} a "
                                f"«{product_name}». Progreso 2026: "
                                f"{_prog_post['linked_campaigns']} de "
                                f"{_prog_post['total_campaigns']} "
                                f"({_prog_post['pct_spend']}% del gasto).",
                                "success",
                            )
                        except Exception:
                            logger.exception(
                                "Error en vinculación masiva de campañas (%d ids)",
                                len(bulk_ids),
                            )
                            banner = (
                                "Error al guardar la vinculación. Intentá de "
                                "nuevo o avisá al administrador.",
                                "danger",
                            )

    elif request.method == "POST":
        campaign_id  = (request.POST.get("campaign_id") or "").strip()
        codigo_str   = (request.POST.get("codigo_producto") or "").strip()
        product_name = (request.POST.get("product_name") or "").strip()
        category     = (request.POST.get("category") or "").strip()

        if not campaign_id or not codigo_str:
            banner = ("Faltan datos: campaign_id o producto.", "danger")
        else:
            try:
                codigo_producto = int(codigo_str)
            except ValueError:
                banner = ("Código de producto inválido.", "danger")
                codigo_producto = None

            if codigo_producto is not None:
                try:
                    from django.utils import timezone as tz
                    MetaCampaignMap.objects.update_or_create(
                        campaign_id=campaign_id,
                        defaults={
                            "codigo_producto": codigo_producto,
                            "product_name": product_name or None,
                            "category": category or None,
                            "linked_by": "manual",
                            "linked_at": tz.now(),
                        },
                    )
                    # Update all tb_meta_ads rows for this campaign_id
                    updated = MetaAds.objects.filter(
                        campaign_id=campaign_id
                    ).update(
                        product=product_name or None,
                        category=category or None,
                    )
                    # Compute progress AFTER the link so the banner reflects
                    # the new state. Scoped to 2026 for consistency with the
                    # pending list and the progress card.
                    _prog_post = _linking_progress(date_from=LINKING_SCOPE_FROM)
                    banner = (
                        f"✓ Vinculada «{product_name}». "
                        f"Progreso 2026: {_prog_post['linked_campaigns']} de "
                        f"{_prog_post['total_campaigns']} campañas "
                        f"({_prog_post['pct_spend']}% del gasto identificado).",
                        "success",
                    )
                except Exception as exc:
                    banner = (f"Error al guardar: {exc}", "danger")

    # --- Progress stats (recomputed after any POST) ---
    # Scoped to 2026: progress card only tracks the campaigns that need linking.
    try:
        progress = _linking_progress(date_from=LINKING_SCOPE_FROM)
    except Exception:
        from decimal import Decimal
        progress = {
            "total_campaigns": 0, "linked_campaigns": 0, "unlinked_campaigns": 0,
            "pct_campaigns": 0, "spend_total": Decimal("0"), "spend_linked": Decimal("0"),
            "spend_unlinked": Decimal("0"), "pct_spend": 0, "pct_spend_unlinked": 100,
            "top10_unlinked_spend": Decimal("0"), "top10_pct_extra": 0,
        }

    # --- Data for the listing ---
    # Campaigns with NO map entry that has codigo_producto set
    linked_ids_with_product = set(
        MetaCampaignMap.objects
        .exclude(codigo_producto__isnull=True)
        .values_list("campaign_id", flat=True)
    )

    # Pending list scoped to 2026: only campaigns with activity in 2026 are
    # shown as needing a link.  Historic 2025-only campaigns stay in the Pauta
    # data but do not appear here (user decision: LINKING_SCOPE_FROM = 2026-01-01).
    # Meta renames campaigns across months, so the same campaign_id can appear
    # with several names — consolidate by campaign_id keeping the most recent name.
    from django.db.models import Max, Sum as DSum
    campaign_qs = (
        MetaAds.objects
        .exclude(campaign_id__isnull=True)
        .filter(report_start__gte=LINKING_SCOPE_FROM)
        .values("campaign_id", "campaign_name", "account_id")
        .annotate(
            recent_spend=DSum("spend"),
            months_seen=Count("month", distinct=True),
            last_seen=Max("report_start"),
        )
        .order_by("-recent_spend")
    )

    by_cid: dict[str, dict] = {}
    for row in campaign_qs:
        cid = row["campaign_id"]
        spend = float(row["recent_spend"] or 0)
        cur = by_cid.get(cid)
        if cur is None:
            by_cid[cid] = {
                "campaign_id":   cid,
                "campaign_name": row["campaign_name"] or cid,
                "account_id":    row["account_id"],
                "recent_spend":  spend,
                "months_seen":   row["months_seen"],
                "_last_seen":    row["last_seen"],
            }
        else:
            cur["recent_spend"] += spend
            cur["months_seen"] += row["months_seen"]
            if row["last_seen"] and (cur["_last_seen"] is None or row["last_seen"] > cur["_last_seen"]):
                cur["campaign_name"] = row["campaign_name"] or cid
                cur["account_id"] = row["account_id"]
                cur["_last_seen"] = row["last_seen"]

    unlinked = []
    linked   = []
    for entry in by_cid.values():
        # last_seen se conserva: el agrupador lo usa para elegir el título
        # representativo (nombre del miembro más reciente).
        entry["last_seen"] = entry.pop("_last_seen", None)
        if entry["campaign_id"] in linked_ids_with_product:
            linked.append(entry)
        else:
            unlinked.append(entry)

    unlinked.sort(key=lambda x: -(x["recent_spend"] or 0))
    linked.sort(key=lambda x: -(x["recent_spend"] or 0))

    products = _get_all_products()

    # --- Agrupación de pendientes + sugerencias de producto ---
    sku_index = {
        (p["sku_producto"] or "").strip().upper(): p
        for p in products if p["sku_producto"]
    }
    name_index = [(_vin_norm_simple(p["nombre_producto"]), p) for p in products]
    groups = _group_pending_campaigns(unlinked, sku_index)
    for g in groups:
        g["suggestion"] = _suggest_product_for_group(g, sku_index, name_index)

    # Count of unlinked for the toolbar badge
    unlinked_count = len(unlinked)

    context = {
        "unlinked":       unlinked,
        "linked":         linked,
        "groups":         groups,
        "groups_count":   len(groups),
        "products":       products,
        "banner":         banner,
        "unlinked_count": unlinked_count,
        "progress":       progress,
        "MAIN_APP_URL":   getattr(settings, "MAIN_APP_URL", ""),
    }
    return render(request, "vincular.html", context)


# ---------------------------------------------------------------------------
# Ads / Crear campaña — campaign creation via Meta Graph API
# ---------------------------------------------------------------------------

# Permission-related error codes from the Meta API
_META_PERMISSION_CODES = {200, 270, 10}
_META_PERMISSION_KEYWORDS = ("ads_management", "permission", "does not have")

CAMPAIGN_OBJECTIVES = [
    ("OUTCOME_LEADS",       "Clientes potenciales"),
    ("OUTCOME_SALES",       "Ventas"),
    ("OUTCOME_TRAFFIC",     "Tráfico"),
    ("OUTCOME_ENGAGEMENT",  "Interacción"),
    ("OUTCOME_AWARENESS",   "Reconocimiento"),
]

MONTHS_ABREV_ES = [
    "ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
    "JUL", "AGO", "SEP", "OCT", "NOV", "DIC",
]


def _meta_permission_error(err_obj: dict) -> bool:
    """Return True if this Meta API error is permission-related."""
    code = err_obj.get("code")
    if code in _META_PERMISSION_CODES:
        return True
    msg = (err_obj.get("message") or "").lower()
    return any(kw in msg for kw in _META_PERMISSION_KEYWORDS)


@ads_admin_required
def ads_crear_campana(request):
    """
    /ads/crear-campana/ — Create a new Meta campaign via the Graph API.
    GET:  Render the form.
    POST: Call the API; on success create a map entry; render result/error.
    """
    token = os.getenv("META_ACCESS_TOKEN", "")
    try:
        from core.management.commands.sync_meta_ads import META_API_VERSION
    except ImportError:
        META_API_VERSION = "v25.0"

    today = datetime.date.today()
    mes_abrev = MONTHS_ABREV_ES[today.month - 1]

    # Active accounts only (account_status = 1)
    accounts = list(
        MetaAccount.objects
        .filter(account_status=1)
        .order_by("name")
        .values("account_id", "name", "currency")
    )
    products = _get_all_products()

    result = None   # dict with success or error info after POST
    form_data = {}  # repopulate form on error

    if request.method == "POST":
        account_id   = (request.POST.get("account_id") or "").strip()
        codigo_str   = (request.POST.get("codigo_producto") or "").strip()
        product_name = (request.POST.get("product_name") or "").strip()
        category     = (request.POST.get("category") or "").strip()
        sku          = (request.POST.get("sku") or "").strip()
        objective    = (request.POST.get("objective") or "").strip()
        camp_name    = (request.POST.get("campaign_name") or "").strip()
        daily_budget = (request.POST.get("daily_budget") or "").strip()

        form_data = {
            "account_id": account_id,
            "codigo_str": codigo_str,
            "product_name": product_name,
            "category": category,
            "sku": sku,
            "objective": objective,
            "campaign_name": camp_name,
            "daily_budget": daily_budget,
        }

        # Basic validation
        errors = []
        if not account_id:
            errors.append("Seleccioná una cuenta publicitaria.")
        if not codigo_str or not product_name:
            errors.append("Seleccioná un producto.")
        if not objective:
            errors.append("Seleccioná un objetivo.")
        if not camp_name:
            errors.append("El nombre de la campaña no puede estar vacío.")

        if errors:
            result = {"type": "validation_error", "errors": errors}
        else:
            # Build API params
            # is_adset_budget_sharing_enabled: required when not using campaign budget (Graph API v25.0+)
            api_params: dict = {
                "name": camp_name,
                "objective": objective,
                "status": "PAUSED",
                "special_ad_categories": "[]",
                "is_adset_budget_sharing_enabled": "false",
                "access_token": token,
            }
            if daily_budget:
                try:
                    from decimal import Decimal as _D
                    budget_minor = int(_D(daily_budget.replace(",", ".")) * 100)
                    api_params["daily_budget"] = str(budget_minor)
                    # With a campaign budget, remove adset sharing flag (not needed)
                    del api_params["is_adset_budget_sharing_enabled"]
                except Exception:
                    pass  # ignore invalid budget — Meta will validate

            act_id = account_id if account_id.startswith("act_") else f"act_{account_id}"
            url = f"https://graph.facebook.com/{META_API_VERSION}/{act_id}/campaigns"
            encoded = urllib.parse.urlencode(api_params).encode("utf-8")

            try:
                req = urllib.request.Request(
                    url,
                    data=encoded,
                    method="POST",
                    headers={"User-Agent": "goberna-dashboard/1.0"},
                )
                with urllib.request.urlopen(req, timeout=30) as resp:
                    resp_data = json.loads(resp.read().decode("utf-8"))

                new_campaign_id = resp_data.get("id") or ""

                # On success: persist map entry
                if new_campaign_id and codigo_str:
                    try:
                        from django.utils import timezone as tz
                        MetaCampaignMap.objects.update_or_create(
                            campaign_id=new_campaign_id,
                            defaults={
                                "codigo_producto": int(codigo_str),
                                "product_name": product_name or None,
                                "category": category or None,
                                "linked_by": "manual",
                                "linked_at": tz.now(),
                            },
                        )
                    except Exception:
                        pass  # Non-fatal: map entry is best-effort

                # Ads Manager deep link
                account_number = account_id.lstrip("act_").lstrip("0") or account_id
                ads_manager_url = (
                    f"https://adsmanager.facebook.com/manage/campaigns"
                    f"?act={account_number}&selected_campaign_ids={new_campaign_id}"
                )

                result = {
                    "type": "success",
                    "campaign_id": new_campaign_id,
                    "campaign_name": camp_name,
                    "ads_manager_url": ads_manager_url,
                }

            except urllib.error.HTTPError as e:
                body = e.read().decode("utf-8", errors="replace")
                try:
                    err_obj = json.loads(body).get("error", {}) or {}
                except Exception:
                    err_obj = {}
                safe_message = err_obj.get("message") or body[:400]
                is_permission = _meta_permission_error(err_obj)
                result = {
                    "type": "api_error",
                    "is_permission": is_permission,
                    "http_code": e.code,
                    "message": safe_message,
                    "error_code": err_obj.get("code"),
                }

            except Exception as exc:
                result = {
                    "type": "api_error",
                    "is_permission": False,
                    "http_code": None,
                    "message": str(exc)[:400],
                    "error_code": None,
                }

    context = {
        "accounts": accounts,
        "products": products,
        "objectives": CAMPAIGN_OBJECTIVES,
        "mes_abrev": mes_abrev,
        "result": result,
        "form_data": form_data,
        "MAIN_APP_URL": getattr(settings, "MAIN_APP_URL", ""),
    }
    return render(request, "crear_campana.html", context)


# ---------------------------------------------------------------------------
# Ads / Embudo — cruce inversión Meta × ventas pagado (funnel view)
# ---------------------------------------------------------------------------

# Country normalization: some ads rows use names not in tb_pais
_ADS_COUNTRY_NORMALIZE = {
    "gran bretaña": "reino unido",
    "bo": "bolivia",
    "co": "colombia",
    "ec": "ecuador",
}


def _normalize_country(name: str | None) -> str:
    """Lower-strip a country name and apply normalization map."""
    if not name:
        return ""
    key = name.strip().lower()
    return _ADS_COUNTRY_NORMALIZE.get(key, key)


def _default_date_range() -> tuple[datetime.date, datetime.date]:
    """Return (desde, hasta) covering the last 3 complete months including current."""
    today = datetime.date.today()
    # 'hasta' = first day of current month
    hasta = datetime.date(today.year, today.month, 1)
    # 'desde' = 3 months back from hasta
    month = hasta.month - 2
    year = hasta.year
    while month <= 0:
        month += 12
        year -= 1
    desde = datetime.date(year, month, 1)
    return desde, hasta


def _parse_date_param(value: str | None) -> datetime.date | None:
    """Parse a YYYY-MM-DD date string or return None."""
    if not value:
        return None
    try:
        return datetime.date.fromisoformat(value.strip())
    except (ValueError, AttributeError):
        return None


def _paid_media_kpis(date_from: datetime.date, date_to: datetime.date) -> dict:
    """
    Compute paid-media funnel KPIs for a date range [date_from, date_to] (both inclusive,
    aligned to month boundaries).

    Returns a dict:
      inversion_usd   – total Meta spend USD (float)
      ventas_usd      – total paid sales USD (float, medio='pagado')
      utilidad_usd    – ventas_usd - inversion_usd
      roas            – ventas_usd / inversion_usd  or None
      margen_pct      – utilidad_usd / ventas_usd * 100  or None
      rows_sin_usd    – count of ads rows excluded due to NULL amount_usd
      por_negocio     – list of {negocio, inversion, ventas, utilidad, margen}
                        grouped by DetalleVenta.producto.codigo_negocio / MetaAds.category
    """
    # ---- Upper bound (exclusive) for date filter ----
    month_after = date_to.month + 1
    year_after = date_to.year
    if month_after > 12:
        month_after = 1
        year_after += 1
    hasta_exclusive = datetime.date(year_after, month_after, 1)

    # ---- Meta spend aggregation ----
    ads_qs = MetaAds.objects.filter(
        report_start__isnull=False,
        report_start__gte=date_from,
        report_start__lte=date_to,
    )
    rows_sin_usd = ads_qs.filter(amount_usd__isnull=True, spend__isnull=False).count()
    agg = ads_qs.aggregate(total_inversion=Sum("amount_usd"))
    inversion_usd = float(agg["total_inversion"] or 0)

    # ---- Sales aggregation (paid only, estado 1/2, medio=pagado) ----
    ventas_qs = (
        Venta.objects
        .filter(estado__in=[1, 2], medio="pagado")
        .annotate(
            tasa_cambio=Coalesce(
                "radio_multiplicador_usado",
                "moneda__radioMultiplicador",
                Value(1),
                output_field=DecimalField(),
            )
        )
        .annotate(
            tasa_final=Case(
                When(tasa_cambio=0, then=Value(1)),
                default=F("tasa_cambio"),
                output_field=DecimalField(),
            )
        )
        .annotate(
            monto_usd=ExpressionWrapper(
                F("monto_total") / F("tasa_final"),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .filter(
            fecha_venta__gte=timezone.make_aware(
                datetime.datetime(date_from.year, date_from.month, date_from.day)
            ),
            fecha_venta__lt=timezone.make_aware(
                datetime.datetime(hasta_exclusive.year, hasta_exclusive.month, hasta_exclusive.day)
            ),
        )
    )
    ventas_agg = ventas_qs.aggregate(total_ventas=Sum("monto_usd"))
    ventas_usd = float(ventas_agg["total_ventas"] or 0)

    utilidad_usd = ventas_usd - inversion_usd
    roas = (ventas_usd / inversion_usd) if inversion_usd > 0 else None
    margen_pct = (utilidad_usd / ventas_usd * 100) if ventas_usd > 0 else None

    # ---- Por negocio: ads grouped by category, sales grouped by negocio ----
    # Ads side: aggregate spend by category
    ads_by_cat = {}
    for row in (
        ads_qs.values("category")
        .annotate(spend_usd=Sum("amount_usd"))
        .order_by()
    ):
        cat = (row.get("category") or "").strip() or "Sin categoría"
        ads_by_cat[cat] = float(row["spend_usd"] or 0)

    # Sales side: aggregate ventas_usd by negocio via DetalleVenta
    sales_by_negocio = {}
    detalles_qs = (
        DetalleVenta.objects
        .filter(venta__in=ventas_qs)
        .annotate(
            tasa_cambio=Coalesce(
                "venta__radio_multiplicador_usado",
                "venta__moneda__radioMultiplicador",
                Value(1),
                output_field=DecimalField(),
            )
        )
        .annotate(
            tasa_final=Case(
                When(tasa_cambio=0, then=Value(1)),
                default=F("tasa_cambio"),
                output_field=DecimalField(),
            )
        )
        .annotate(
            monto_usd=ExpressionWrapper(
                F("precio_total") / F("tasa_final"),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .values("producto__codigo_negocio__nombre_negocio")
        .annotate(ventas_usd=Sum("monto_usd"))
        .order_by()
    )
    for row in detalles_qs:
        neg = (row.get("producto__codigo_negocio__nombre_negocio") or "").strip() or "Sin negocio"
        sales_by_negocio[neg] = float(row["ventas_usd"] or 0)

    # Merge: union of keys from both sides
    all_keys = set(ads_by_cat.keys()) | set(sales_by_negocio.keys())
    por_negocio = []
    for key in sorted(all_keys):
        inv = ads_by_cat.get(key, 0.0)
        ven = sales_by_negocio.get(key, 0.0)
        util = ven - inv
        marg = (util / ven * 100) if ven > 0 else None
        por_negocio.append({
            "negocio": key,
            "inversion": inv,
            "ventas": ven,
            "utilidad": util,
            "margen": marg,
        })
    # Sort by utilidad desc
    por_negocio.sort(key=lambda x: -x["utilidad"])

    return {
        "inversion_usd": inversion_usd,
        "ventas_usd": ventas_usd,
        "utilidad_usd": utilidad_usd,
        "roas": roas,
        "margen_pct": margen_pct,
        "rows_sin_usd": rows_sin_usd,
        "por_negocio": por_negocio,
    }


def _roas_por_producto(date_from: datetime.date, date_to: datetime.date) -> list[dict]:
    """
    ROAS per product for [date_from, date_to]: Meta spend (amount_usd) crossed
    with paid sales (estado 1/2, medio='pagado') converted to USD.

    Join key: codigo_producto via MetaCampaignMap when available, else exact
    product-name match against the Producto catalog. Spend with no resolvable
    product is bucketed under "Sin producto vinculado".

    Only products with spend in the period are returned (the home dashboard
    already covers the global picture; here the question is campaign ROI).
    """
    month_after = date_to.month + 1
    year_after = date_to.year
    if month_after > 12:
        month_after = 1
        year_after += 1
    hasta_exclusive = datetime.date(year_after, month_after, 1)

    # ---- Lookups: campaign -> codigo, nombre -> codigo, codigo -> nombre ----
    cmap = {
        m.campaign_id: m.codigo_producto
        for m in MetaCampaignMap.objects.filter(codigo_producto__isnull=False)
    }
    name_to_codigo = {}
    codigo_to_name = {}
    for cod, nom in Producto.objects.values_list("codigo_producto", "nombre_producto"):
        codigo_to_name[cod] = nom
        if nom:
            name_to_codigo[nom.strip()] = cod

    # ---- Meta side: spend per codigo_producto ----
    # BOB-billed accounts have amount_usd=NULL; convert spend (BOB) -> USD
    # on the fly using the moneda BOB ratio (codigo_moneda=3), with a
    # hardcoded fallback to the current rate if not configured.
    bob_ratio = Moneda.objects.filter(pk=3).values_list(
        "radioMultiplicador", flat=True
    ).first()
    bob_ratio = bob_ratio or Decimal("9.07")

    spend_by_codigo: dict = {}
    spend_sin_producto = 0.0
    ads_rows = (
        MetaAds.objects
        .filter(
            report_start__gte=date_from,
            report_start__lte=date_to,
        )
        .annotate(
            gasto_usd=Case(
                When(amount_usd__isnull=False, then=F("amount_usd")),
                When(
                    amount_usd__isnull=True, account_currency="BOB",
                    then=ExpressionWrapper(
                        F("spend") / Value(bob_ratio, output_field=DecimalField()),
                        output_field=DecimalField(max_digits=14, decimal_places=4),
                    ),
                ),
                default=Value(0),
                output_field=DecimalField(max_digits=14, decimal_places=4),
            )
        )
        .values("campaign_id", "product")
        .annotate(spend_usd=Sum("gasto_usd"))
        .order_by()
    )
    for row in ads_rows:
        codigo = cmap.get(row["campaign_id"])
        if codigo is None and row["product"]:
            codigo = name_to_codigo.get(row["product"].strip())
        usd = float(row["spend_usd"] or 0)
        if codigo is None:
            spend_sin_producto += usd
        else:
            spend_by_codigo[codigo] = spend_by_codigo.get(codigo, 0.0) + usd

    # ---- Sales side: paid sales USD per codigo_producto ----
    ventas_by_codigo: dict = {}
    detalles_qs = (
        DetalleVenta.objects
        .filter(
            venta__estado__in=[1, 2],
            venta__medio="pagado",
            venta__fecha_venta__gte=timezone.make_aware(
                datetime.datetime(date_from.year, date_from.month, date_from.day)
            ),
            venta__fecha_venta__lt=timezone.make_aware(
                datetime.datetime(hasta_exclusive.year, hasta_exclusive.month, hasta_exclusive.day)
            ),
        )
        .annotate(
            tasa_cambio=Coalesce(
                "venta__radio_multiplicador_usado",
                "venta__moneda__radioMultiplicador",
                Value(1),
                output_field=DecimalField(),
            )
        )
        .annotate(
            tasa_final=Case(
                When(tasa_cambio=0, then=Value(1)),
                default=F("tasa_cambio"),
                output_field=DecimalField(),
            )
        )
        .annotate(
            monto_usd=ExpressionWrapper(
                F("precio_total") / F("tasa_final"),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .values("producto__codigo_producto")
        .annotate(total_usd=Sum("monto_usd"))
        .order_by()
    )
    for row in detalles_qs:
        cod = row["producto__codigo_producto"]
        if cod is not None:
            ventas_by_codigo[cod] = float(row["total_usd"] or 0)

    # ---- Merge: only products with spend in the period ----
    out = []
    for codigo, inv in spend_by_codigo.items():
        ven = ventas_by_codigo.get(codigo, 0.0)
        out.append({
            "codigo": codigo,
            "producto": codigo_to_name.get(codigo) or f"Producto {codigo}",
            "inversion": round(inv, 2),
            "ventas": round(ven, 2),
            "utilidad": round(ven - inv, 2),
            "roas": round(ven / inv, 2) if inv > 0 else None,
        })
    if spend_sin_producto > 0:
        out.append({
            "codigo": None,
            "producto": "Sin producto vinculado",
            "inversion": round(spend_sin_producto, 2),
            "ventas": None,
            "utilidad": None,
            "roas": None,
        })
    out.sort(key=lambda r: -r["inversion"])
    return out


def _roas_por_producto_pais(date_from: datetime.date, date_to: datetime.date) -> list[dict]:
    """
    ROAS per (producto, pais) for [date_from, date_to]: Meta spend (amount_usd,
    with on-the-fly BOB->USD conversion) crossed with paid sales (estado 1/2,
    medio='pagado') converted to USD, both broken down by country.

    Product resolution follows the same chain as _roas_por_producto:
    MetaCampaignMap.codigo_producto -> fallback exact product-name match ->
    "Sin producto vinculado" (codigo=None), kept broken down by country.

    Unlike _roas_por_producto, rows with inversion_usd == 0 but ventas_usd > 0
    are also included (sales without matching ad spend in that country).
    """
    month_after = date_to.month + 1
    year_after = date_to.year
    if month_after > 12:
        month_after = 1
        year_after += 1
    hasta_exclusive = datetime.date(year_after, month_after, 1)

    # ---- Lookups: campaign -> codigo, nombre -> codigo, codigo -> nombre ----
    cmap = {
        m.campaign_id: m.codigo_producto
        for m in MetaCampaignMap.objects.filter(codigo_producto__isnull=False)
    }
    name_to_codigo = {}
    codigo_to_name = {}
    for cod, nom in Producto.objects.values_list("codigo_producto", "nombre_producto"):
        codigo_to_name[cod] = nom
        if nom:
            name_to_codigo[nom.strip()] = cod

    # ---- Meta side: spend per (codigo_producto, pais) ----
    bob_ratio = Moneda.objects.filter(pk=3).values_list(
        "radioMultiplicador", flat=True
    ).first()
    bob_ratio = bob_ratio or Decimal("9.07")

    inversion_by_key: dict = {}
    ads_rows = (
        MetaAds.objects
        .filter(
            report_start__gte=date_from,
            report_start__lte=date_to,
        )
        .annotate(
            gasto_usd=Case(
                When(amount_usd__isnull=False, then=F("amount_usd")),
                When(
                    amount_usd__isnull=True, account_currency="BOB",
                    then=ExpressionWrapper(
                        F("spend") / Value(bob_ratio, output_field=DecimalField()),
                        output_field=DecimalField(max_digits=14, decimal_places=4),
                    ),
                ),
                default=Value(0),
                output_field=DecimalField(max_digits=14, decimal_places=4),
            )
        )
        .values("campaign_id", "campaign_name", "product", "paid_country", "country")
        .annotate(spend_usd=Sum("gasto_usd"))
        .order_by()
    )
    # Para filas "Sin producto vinculado" (codigo is None): gasto por campaña
    # individual dentro de cada (None, pais), para que el frontend pueda
    # mostrar "qué campañas componen este monto" (botón "Ver campañas").
    campanas_by_key: dict = {}
    for row in ads_rows:
        codigo = cmap.get(row["campaign_id"])
        if codigo is None and row["product"]:
            codigo = name_to_codigo.get(row["product"].strip())

        pais = (row["paid_country"] or "").strip() or (row["country"] or "").strip() or "Sin país"

        usd = float(row["spend_usd"] or 0)
        key = (codigo, pais)
        inversion_by_key[key] = inversion_by_key.get(key, 0.0) + usd

        if codigo is None and usd > 0:
            cname = (row["campaign_name"] or row["campaign_id"] or "").strip()
            if cname:
                by_campaign = campanas_by_key.setdefault(key, {})
                by_campaign[cname] = by_campaign.get(cname, 0.0) + usd

    # ---- Sales side: paid sales USD + count per (codigo_producto, pais) ----
    ventas_usd_by_key: dict = {}
    ventas_count_by_key: dict = {}
    detalles_qs = (
        DetalleVenta.objects
        .filter(
            venta__estado__in=[1, 2],
            venta__medio="pagado",
            venta__fecha_venta__gte=timezone.make_aware(
                datetime.datetime(date_from.year, date_from.month, date_from.day)
            ),
            venta__fecha_venta__lt=timezone.make_aware(
                datetime.datetime(hasta_exclusive.year, hasta_exclusive.month, hasta_exclusive.day)
            ),
        )
        .annotate(
            tasa_cambio=Coalesce(
                "venta__radio_multiplicador_usado",
                "venta__moneda__radioMultiplicador",
                Value(1),
                output_field=DecimalField(),
            )
        )
        .annotate(
            tasa_final=Case(
                When(tasa_cambio=0, then=Value(1)),
                default=F("tasa_cambio"),
                output_field=DecimalField(),
            )
        )
        .annotate(
            monto_usd=ExpressionWrapper(
                F("precio_total") / F("tasa_final"),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
        .values(
            "producto__codigo_producto",
            "venta__pais__nombre",
            "venta__cliente__pais__nombre",
        )
        .annotate(total_usd=Sum("monto_usd"), total_count=Count("id"))
        .order_by()
    )
    for row in detalles_qs:
        cod = row["producto__codigo_producto"]
        pais = (
            (row["venta__pais__nombre"] or "").strip()
            or (row["venta__cliente__pais__nombre"] or "").strip()
            or "Sin país"
        )
        key = (cod, pais)
        ventas_usd_by_key[key] = ventas_usd_by_key.get(key, 0.0) + float(row["total_usd"] or 0)
        ventas_count_by_key[key] = ventas_count_by_key.get(key, 0) + (row["total_count"] or 0)

    # ---- Merge: union of all (codigo_producto, pais) keys from both sides ----
    all_keys = set(inversion_by_key.keys()) | set(ventas_usd_by_key.keys()) | set(ventas_count_by_key.keys())

    out = []
    for codigo, pais in all_keys:
        inv = inversion_by_key.get((codigo, pais), 0.0)
        ven = ventas_usd_by_key.get((codigo, pais), 0.0)
        count = ventas_count_by_key.get((codigo, pais), 0)
        out.append({
            "codigo_producto": codigo,
            "producto": codigo_to_name.get(codigo) or "Sin producto vinculado",
            "pais": pais,
            "ventas_count": count,
            "inversion_usd": round(inv, 2),
            "ventas_usd": round(ven, 2),
            "utilidad_usd": round(ven - inv, 2),
            "roas": round(ven / inv, 2) if inv > 0 else None,
            "campanas": sorted(
                (
                    {"nombre": nombre, "gasto_usd": round(gasto, 2)}
                    for nombre, gasto in campanas_by_key.get((codigo, pais), {}).items()
                ),
                key=lambda c: -c["gasto_usd"],
            ) if codigo is None else [],
        })

    out.sort(key=lambda r: -r["inversion_usd"])
    return out


@ads_admin_required
def pautas_ventas_cursos(request):
    """
    Página "Pautas y Ventas por Cursos".
    - GET (browser): renderiza pautas_cursos.html
    - GET con X-Requested-With: XMLHttpRequest (o ?data=1): devuelve JSON
      {"date_from": "...", "date_to": "...", "rows": [...]} usando
      _roas_por_producto_pais (misma función ya validada en ads_dashboard).

    Parámetros GET opcionales: date_from / date_to (ISO YYYY-MM-DD).
    Por defecto: año actual completo (1 enero - hoy), igual al patrón de
    ads_dashboard (?roas=1).
    """
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    today = datetime.date.today()
    try:
        r_from = datetime.date.fromisoformat(date_from) if date_from else datetime.date(today.year, 1, 1)
        r_to = datetime.date.fromisoformat(date_to) if date_to else today
    except ValueError:
        return JsonResponse({"error": "fecha inválida"}, status=400)

    is_ajax = (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or request.GET.get("data", "").strip() == "1"
    )

    if is_ajax:
        return JsonResponse({
            "date_from": r_from.isoformat(),
            "date_to": r_to.isoformat(),
            "rows": _roas_por_producto_pais(r_from, r_to),
        })

    return render(request, "pautas_cursos.html", {
        "date_from": r_from.isoformat(),
        "date_to": r_to.isoformat(),
    })

