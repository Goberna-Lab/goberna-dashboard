"""
Management command: import_meta_ads
Loads the "Pauta" sheet from a Meta Ads Manager Excel export into tb_meta_ads.

Usage:
    python manage.py import_meta_ads <path_to_xlsx> [--sheet=Pauta] [--replace] [--dry-run]

Options:
    --sheet     Name of the Excel sheet to read (default: Pauta).
    --replace   Delete all source='excel' rows + re-insert inside a single transaction.
                Rows with source='api' (written by sync_meta_ads) are NEVER touched.
                Without this flag the command does an append (useful for incremental loads).
    --dry-run   Parse, validate, and report — but write nothing to the DB.

Notes:
- campaign_id and account_id are stored as VARCHAR to avoid float64 precision loss.
- The command bootstraps the FULL schema (shared with sync_meta_ads via
  _meta_ads_schema): CREATE TABLE IF NOT EXISTS + additive ALTERs for older
  tables — no Meta API token needed (Django never does DDL because MetaAds
  is managed=False).
- Row-level validation is fail-soft: bad rows are logged and skipped, the rest
  are imported.
- Uses openpyxl (read_only=True, data_only=True) — pandas is NOT a dependency.
"""

import datetime
import math
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction

from core.models import MetaAds

from ._meta_ads_schema import ensure_full_schema

# ---------------------------------------------------------------------------
# Column header → model field mapping  (Spanish header → Python field name)
# ---------------------------------------------------------------------------
COLUMN_MAP = {
    "Nombre de la Campaña":     "campaign_name",
    "ID de la Campaña":         "campaign_id",
    "ID de la Cuenta":          "account_id",
    "Producto":                 "product",
    "Categoría":                "category",
    "Mes":                      "month",
    "País Pagado":              "paid_country",
    "País":                     "country",
    "Entrega":                  "delivery",
    "Resultados":               "results",
    "Indicador de resultados":  "result_indicator",
    "Alcance":                  "reach",
    "Impresiones":              "impressions",
    "Link clicks":              "link_clicks",
    "Costo por Resultados":     "cost_per_result",
    "Importe Gastado":          "spend",
    "Monto Dólares":            "amount_usd",
    "Inicio":                   "start_date",
    "Fin":                      "end_date",
    "Inicio del Informe":       "report_start",
    "Fin del Informe":          "report_end",
}

# ---------------------------------------------------------------------------
# Normalization alias maps
# ---------------------------------------------------------------------------
COUNTRY_ALIASES = {
    "perú nuevo":              "Perú",
    "peru nuevo":              "Perú",
    "república dominicana":    "República Dominicana",
    "republica dominicana":    "República Dominicana",
    "republica":               "República",
}

MONTH_ALIASES = {
    "setiembre": "Septiembre",
    "SETIEMBRE": "Septiembre",
    "Setiembre": "Septiembre",
}

# DDL for tb_meta_ads lives in _meta_ads_schema (shared with sync_meta_ads).


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _is_empty(v) -> bool:
    """Return True for None, empty string, 'N/A', and float NaN."""
    if v is None:
        return True
    s = str(v).strip()
    if s == "" or s.upper() in ("N/A", "NA", "NONE", "NULL"):
        return True
    try:
        f = float(s)
        return math.isnan(f)
    except (ValueError, TypeError):
        return False


def _clean_id(v) -> tuple[str | None, bool]:
    """
    Coerce a cell value to a clean integer string for campaign_id / account_id.
    Returns (str_value_or_None, was_float_coerced).

    If the cell value is already a string with only digits it is returned as-is.
    If it is a float (openpyxl may return 2.385e+16 for IDs stored as numbers)
    we convert via int() — valid only when the float has no fractional part AND
    the integer is safely representable.  If precision loss is suspected we flag it.
    """
    if _is_empty(v):
        return None, False

    s = str(v).strip()

    # Already a clean integer string (17-18 digits typical for Meta IDs)
    if s.isdigit():
        return s, False

    # Float path
    try:
        f = float(s)
        if math.isnan(f) or math.isinf(f):
            return None, False
        # Check if fractional part exists
        if f != int(f):
            # Has fractional — treat as string anyway; odd for an ID
            return s, True
        i = int(f)
        result = str(i)
        # Warn if the string looks like it may have lost precision:
        # float64 can represent ints up to 2^53 exactly.
        was_coerced = True  # it was originally a float cell
        return result, was_coerced
    except (ValueError, TypeError):
        # Return raw string — let downstream handle it
        return s, False


def _norm_country(v) -> str | None:
    """Normalize country/paid_country: trim, alias lookup, title-case."""
    if _is_empty(v):
        return None
    s = str(v).strip()
    lower = s.lower()
    if lower in COUNTRY_ALIASES:
        return COUNTRY_ALIASES[lower]
    return s.title()


def _norm_month(v) -> str | None:
    """Normalize month: trim, alias lookup, title-case."""
    if _is_empty(v):
        return None
    s = str(v).strip()
    if s in MONTH_ALIASES:
        return MONTH_ALIASES[s]
    lower = s.lower()
    # Build a case-insensitive alias lookup
    for alias_key, alias_val in MONTH_ALIASES.items():
        if alias_key.lower() == lower:
            return alias_val
    return s.title()


def _to_int(v) -> int | None:
    """Convert a cell value to int, returning None for empty/unparseable."""
    if _is_empty(v):
        return None
    try:
        f = float(str(v).strip())
        if math.isnan(f):
            return None
        return int(f)
    except (ValueError, TypeError):
        return None


def _to_decimal(v) -> Decimal | None:
    """Convert a cell value to Decimal(12,2), returning None for empty/unparseable."""
    if _is_empty(v):
        return None
    try:
        s = str(v).strip().replace(",", ".")
        d = Decimal(s)
        return d.quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _to_date(v) -> datetime.date | None:
    """
    Parse a cell value to a date.
    Accepts: datetime.date, datetime.datetime, ISO string, dd/mm/yyyy string.
    Returns None for unparseable values.
    """
    if _is_empty(v):
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    # Try ISO format first
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = (
        "Load the Meta Ads 'Pauta' sheet from an Excel (.xlsx) file into "
        "tb_meta_ads. Creates the table if it doesn't exist."
    )

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Path to the .xlsx file")
        parser.add_argument(
            "--sheet",
            type=str,
            default="Pauta",
            help="Sheet name to read (default: Pauta)",
        )
        parser.add_argument(
            "--replace",
            action="store_true",
            default=False,
            help=(
                "Delete all source='excel' rows before inserting (full excel refresh). "
                "API-synced rows (source='api') are never touched. Without this, rows are appended."
            ),
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            default=False,
            help="Parse and validate only — do not write to the DB.",
        )

    def handle(self, *args, **options):
        path = options["path"]
        sheet_name = options["sheet"]
        replace = options["replace"]
        dry_run = options["dry_run"]

        # ------------------------------------------------------------------
        # 1. Emit DDL (full schema: CREATE TABLE IF NOT EXISTS + additive ALTERs)
        # ------------------------------------------------------------------
        if not dry_run:
            self.stdout.write("Verificando / creando tabla tb_meta_ads...")
            try:
                added = ensure_full_schema(connection)
                if added:
                    self.stdout.write(self.style.SUCCESS(f"  schema actualizado: {', '.join(added)}"))
                else:
                    self.stdout.write(self.style.SUCCESS("  tabla OK"))
            except Exception as exc:
                raise CommandError(
                    f"No se pudo crear/actualizar la tabla tb_meta_ads. "
                    f"Verificar permisos CREATE/ALTER en la DB.\nError: {exc}"
                )

        # ------------------------------------------------------------------
        # 2. Load workbook
        # ------------------------------------------------------------------
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl no está instalado. Ejecutar: pip install openpyxl")

        self.stdout.write(f"Abriendo {path} (hoja: {sheet_name})...")
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Archivo no encontrado: {path}")
        except Exception as exc:
            raise CommandError(f"Error al abrir el archivo: {exc}")

        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise CommandError(
                f"Hoja '{sheet_name}' no encontrada. "
                f"Hojas disponibles: {available}"
            )

        ws = wb[sheet_name]

        # ------------------------------------------------------------------
        # 3. Parse headers
        # ------------------------------------------------------------------
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise CommandError("La hoja está vacía.")

        headers = [str(h).strip() if h is not None else "" for h in header_row]

        # Build index → field_name mapping (ignoring unknown columns)
        col_index_map: dict[int, str] = {}
        for idx, h in enumerate(headers):
            if h in COLUMN_MAP:
                col_index_map[idx] = COLUMN_MAP[h]

        missing = set(COLUMN_MAP.keys()) - set(headers)
        if missing:
            self.stdout.write(
                self.style.WARNING(
                    f"  Columnas esperadas no encontradas en el Excel: {', '.join(sorted(missing))}"
                )
            )

        # ------------------------------------------------------------------
        # 4. Parse rows
        # ------------------------------------------------------------------
        objects: list[MetaAds] = []
        errors: list[dict] = []
        normalization_changes = 0
        float_id_warnings = 0
        rows_read = 0
        rows_skipped = 0

        for row_num, row in enumerate(rows_iter, start=2):  # data starts at row 2
            rows_read += 1
            row_data: dict[str, object] = {}
            for idx, field in col_index_map.items():
                if idx < len(row):
                    row_data[field] = row[idx]
                else:
                    row_data[field] = None

            # --- Required fields ---
            raw_campaign_id = row_data.get("campaign_id")
            raw_campaign_name = row_data.get("campaign_name")

            if _is_empty(raw_campaign_id) or _is_empty(raw_campaign_name):
                reason = (
                    "missing campaign_id" if _is_empty(raw_campaign_id)
                    else "missing campaign_name"
                )
                errors.append({
                    "row": row_num,
                    "reason": reason,
                    "value": raw_campaign_id or raw_campaign_name,
                })
                rows_skipped += 1
                continue

            # --- ID fields ---
            campaign_id, c_coerced = _clean_id(raw_campaign_id)
            account_id_raw = row_data.get("account_id")
            account_id, a_coerced = _clean_id(account_id_raw)

            if c_coerced or a_coerced:
                float_id_warnings += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"  Fila {row_num}: ID convertido de float a string "
                        f"(campaign_id={campaign_id}, account_id={account_id}). "
                        f"Verificar precisión vs fuente Meta."
                    )
                )

            # --- Country normalization ---
            raw_paid_country = row_data.get("paid_country")
            raw_country = row_data.get("country")
            paid_country = _norm_country(raw_paid_country)
            country = _norm_country(raw_country)
            if paid_country and raw_paid_country and str(raw_paid_country).strip() != paid_country:
                normalization_changes += 1
            if country and raw_country and str(raw_country).strip() != country:
                normalization_changes += 1

            # --- Month normalization ---
            raw_month = row_data.get("month")
            month = _norm_month(raw_month)
            if month and raw_month and str(raw_month).strip() != month:
                normalization_changes += 1

            # --- Date fields ---
            date_parse_errors = []
            start_date = _to_date(row_data.get("start_date"))
            if row_data.get("start_date") and not _is_empty(row_data.get("start_date")) and start_date is None:
                date_parse_errors.append(f"start_date={row_data.get('start_date')!r}")

            end_date = _to_date(row_data.get("end_date"))
            if row_data.get("end_date") and not _is_empty(row_data.get("end_date")) and end_date is None:
                date_parse_errors.append(f"end_date={row_data.get('end_date')!r}")

            report_start = _to_date(row_data.get("report_start"))
            if row_data.get("report_start") and not _is_empty(row_data.get("report_start")) and report_start is None:
                date_parse_errors.append(f"report_start={row_data.get('report_start')!r}")

            report_end = _to_date(row_data.get("report_end"))
            if row_data.get("report_end") and not _is_empty(row_data.get("report_end")) and report_end is None:
                date_parse_errors.append(f"report_end={row_data.get('report_end')!r}")

            if date_parse_errors:
                errors.append({
                    "row": row_num,
                    "reason": "date_parse_failure",
                    "value": "; ".join(date_parse_errors),
                })

            # --- Build MetaAds instance ---
            obj = MetaAds(
                campaign_name=str(raw_campaign_name).strip() if not _is_empty(raw_campaign_name) else None,
                campaign_id=campaign_id,
                account_id=account_id,
                product=str(row_data.get("product")).strip() if not _is_empty(row_data.get("product")) else None,
                category=str(row_data.get("category")).strip() if not _is_empty(row_data.get("category")) else None,
                month=month,
                paid_country=paid_country,
                country=country,
                delivery=str(row_data.get("delivery")).strip() if not _is_empty(row_data.get("delivery")) else None,
                results=_to_int(row_data.get("results")),
                result_indicator=str(row_data.get("result_indicator")).strip() if not _is_empty(row_data.get("result_indicator")) else None,
                reach=_to_int(row_data.get("reach")),
                impressions=_to_int(row_data.get("impressions")),
                link_clicks=_to_int(row_data.get("link_clicks")),
                cost_per_result=_to_decimal(row_data.get("cost_per_result")),
                spend=_to_decimal(row_data.get("spend")),
                amount_usd=_to_decimal(row_data.get("amount_usd")),
                start_date=start_date,
                end_date=end_date,
                report_start=report_start,
                report_end=report_end,
            )
            objects.append(obj)

        wb.close()

        # ------------------------------------------------------------------
        # 5. Summary (always printed, even in dry-run)
        # ------------------------------------------------------------------
        rows_valid = rows_read - rows_skipped

        self.stdout.write("")
        self.stdout.write("=" * 60)
        self.stdout.write(f"  Filas leídas         : {rows_read}")
        self.stdout.write(f"  Filas válidas        : {rows_valid}")
        self.stdout.write(f"  Filas saltadas       : {rows_skipped}")
        self.stdout.write(f"  Cambios normalización: {normalization_changes}")
        self.stdout.write(f"  Advertencias IDs float: {float_id_warnings}")
        self.stdout.write(f"  Errores (acumulados) : {len(errors)}")

        if errors:
            self.stdout.write("")
            self.stdout.write("  Top errores (máx 10):")
            for e in errors[:10]:
                self.stdout.write(f"    Fila {e['row']:>5}: {e['reason']} | {e['value']}")

        self.stdout.write("=" * 60)

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY-RUN activo — ningún dato fue escrito a la DB."))
            return

        # ------------------------------------------------------------------
        # 6. Write to DB
        # ------------------------------------------------------------------
        self.stdout.write(f"Escribiendo {len(objects)} filas en tb_meta_ads...")

        try:
            with transaction.atomic():
                if replace:
                    # Only excel rows: legacy rows default to source='excel' (the
                    # column is guaranteed by ensure_full_schema in step 1).
                    # API-synced rows (source='api') belong to sync_meta_ads.
                    deleted = MetaAds.objects.filter(source="excel").delete()
                    self.stdout.write(
                        f"  --replace: {deleted[0]} filas excel anteriores eliminadas "
                        f"(filas source='api' intactas)."
                    )

                MetaAds.objects.bulk_create(objects, batch_size=500)

            self.stdout.write(
                self.style.SUCCESS(
                    f"  Importación completada: {len(objects)} filas insertadas."
                )
            )
        except Exception as exc:
            raise CommandError(
                f"Error durante la escritura en la DB (rollback completo): {exc}"
            )
